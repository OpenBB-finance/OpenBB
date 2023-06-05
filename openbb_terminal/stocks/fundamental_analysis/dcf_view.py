""" DCF View """
__docformat__ = "numpy"

import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

import financedatabase as fd
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from sklearn.linear_model import LinearRegression

from openbb_terminal.decorators import log_start_end
from openbb_terminal.helper_funcs import get_rf
from openbb_terminal.rich_config import console
from openbb_terminal.stocks.fundamental_analysis import dcf_model, dcf_static

# pylint: disable=C0302,too-many-arguments,too-many-boolean-expressions


logger = logging.getLogger(__name__)


class CreateExcelFA:
    @log_start_end(log=logger)
    def __init__(
        self,
        symbol: str,
        beta: float,
        audit: bool = False,
        ratios: bool = True,
        len_pred: int = 10,
        max_similars: int = 3,
        no_filter: bool = False,
        growth: bool = False,
    ):
        """
        Creates a detailed DCF for a given company

        Parameters
        ----------
        symbol : str
            The ticker symbol to create a DCF for
        audit : bool
            Whether or not to show that the balance sheet and income statement tie-out
        ratios : bool
            Whether to show ratios for the company and for similar companies
        len_pred : int
            The number of years to make predictions for before assuming a terminal value
        max_similars : int
            The maximum number of similar companies to show, will be less if there are not enough
        no_filter : bool
            Disable filtering of similar companies to being in the same market cap category
        growth : bool
            When true this turns the revenue assumption from linear regression to percentage growth
        """
        self.info: Dict[str, Any] = {
            "len_data": 0,
            "len_pred": len_pred,
            "max_similars": max_similars,
            "rounding": 0,
            "symbol": symbol,
            "audit": audit,
            "ratios": ratios,
            "no_filter": no_filter,
        }
        self.letter: int = 0
        self.growth = growth
        self.starts: Dict[str, int] = {"IS": 4, "BS": 18, "CF": 47}
        self.wb: Workbook = Workbook()
        self.ws: Dict[int, Any] = {
            1: self.wb.active,
            2: self.wb.create_sheet("Free Cash Flows"),
            3: self.wb.create_sheet("Explanations"),
        }
        self.df: Dict[str, pd.DataFrame] = {
            "BS": self.get_data("BS", self.starts["BS"], False),
            "IS": self.get_data("IS", self.starts["IS"], True),
            "CF": self.get_data("CF", self.starts["CF"], False),
        }

        df_not_created = (
            (not isinstance(self.df["BS"], pd.DataFrame) and not self.df["BS"])
            and (not isinstance(self.df["IS"], pd.DataFrame) and not self.df["IS"])
            and (not isinstance(self.df["CF"], pd.DataFrame) and not self.df["CF"])
        )
        equities_database = fd.Equities()
        self.data: Dict[str, Any] = (
            {
                "now": datetime.now().strftime("%Y-%m-%d"),
                "info": {
                    "country": equities_database.data.loc[self.info["symbol"]][
                        "country"
                    ],
                    "sector": equities_database.data.loc[self.info["symbol"]]["sector"],
                    "industry_group": equities_database.data.loc[self.info["symbol"]][
                        "industry_group"
                    ],
                    "industry": equities_database.data.loc[self.info["symbol"]][
                        "industry"
                    ],
                    "beta": beta,
                },
                "t_bill": get_rf(),
                "r_ff": dcf_model.get_fama_coe(self.info["symbol"]),
                "f_info": yf.Ticker(symbol).fast_info,
            }
            if not df_not_created
            else {}
        )

        if df_not_created:
            console.print(
                "It was not possible to create a DCF for this company.\n"
                "Please consider that non US-listed companies are not supported."
            )

    @log_start_end(log=logger)
    def create_workbook(self):
        self.ws[1].title = "Financials"
        self.ws[1].column_dimensions["A"].width = 25
        self.ws[2].column_dimensions["A"].width = 22
        for column in dcf_static.letters[1:21]:
            self.ws[1].column_dimensions[column].width = 14
        for column in dcf_static.letters[1:21]:
            self.ws[2].column_dimensions[column].width = 14
        self.ws[3].column_dimensions["A"].width = 3

        for _, value in self.ws.items():
            self.create_header(value)

        self.df["BS"], self.df["IS"], self.df["CF"] = dcf_model.clean_dataframes(
            self.df["BS"], self.df["IS"], self.df["CF"]
        )
        self.add_estimates()
        self.create_dcf()
        if self.info["ratios"]:
            self.add_ratios()
        if self.info["audit"]:
            self.run_audit()

        i = 0
        while True:
            file_name = f"{self.data['now'].replace('-', '')}_dcf_{self.info['symbol']}"
            path = dcf_model.generate_path(i, file_name)
            path.parent.mkdir(parents=True, exist_ok=True)

            if not path.is_file():
                self.wb.save(path)
                console.print(
                    f"Find the Discounted Cash Flow (DCF) Analysis of {self.info['symbol']} here: {path}"
                )
                break
            i += 1

    @log_start_end(log=logger)
    def get_data(self, statement: str, row: int, header: bool) -> pd.DataFrame:
        symbol = self.info["symbol"]
        df, rounding, _ = dcf_model.create_dataframe(symbol, statement)
        if df.empty:
            raise ValueError(
                f"Could not generate a dataframe for the ticker `{symbol}` and statement `{statement}`."
            )
        self.info["rounding"] = rounding
        if not self.info["len_data"]:
            self.info["len_data"] = len(df.columns)

        self.ws[1][f"A{row}"] = dcf_static.statement_titles[statement]
        self.ws[1][f"A{row}"].font = dcf_static.bold_font

        rowI = row + 1
        names = df.index.values.tolist()

        for name in names:
            self.ws[1][f"A{rowI}"] = name
            if name in dcf_static.sum_rows:
                length = self.info["len_data"] + (
                    self.info["len_pred"] if statement != "CF" else 0
                )
                for i in range(length):
                    if statement == "CF" and name == "Net Income":
                        pass
                    else:
                        self.ws[1][
                            f"{dcf_static.letters[i+1]}{rowI}"
                        ].font = dcf_static.bold_font
                        self.ws[1][
                            f"{dcf_static.letters[i+1]}{rowI}"
                        ].border = dcf_static.thin_border_top
            rowI += 1

        column = 1

        for key, value in df.items():
            rowI = row
            if header:
                dcf_model.set_cell(
                    self.ws[1],
                    f"{dcf_static.letters[column]}{rowI}",
                    float(key),
                    font=dcf_static.bold_font,
                )
            for item in value:
                rowI += 1
                try:
                    m = 0 if item is None else float(item.replace(",", ""))
                except ValueError:
                    m = 0
                dcf_model.set_cell(
                    self.ws[1],
                    f"{dcf_static.letters[column]}{rowI}",
                    m,
                    num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
                )
            column += 1

        return df

    @log_start_end(log=logger)
    def add_estimates(self):
        last_year = self.df["BS"].columns[-1]  # Replace with columns in DF
        col = self.info["len_data"] + 1
        for i in range(self.info["len_pred"]):
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+i]}4",
                int(last_year) + 1 + i,
                font=dcf_static.bold_font,
            )

        for i in range(41):
            col = self.info["len_pred"] + self.info["len_data"] + 3
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col]}{3+i}",
                fill=dcf_static.green_bg,
                border=dcf_static.thin_border_nr,
            )
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+1]}{3+i}",
                fill=dcf_static.green_bg,
                border=dcf_static.thin_border_nl,
            )

        dcf_model.set_cell(
            self.ws[1],
            f"{dcf_static.letters[col]}3",
            "Linear model",
            alignment=dcf_static.center,
        )
        self.ws[1].merge_cells(
            f"{dcf_static.letters[col]}3:{dcf_static.letters[col+1]}3"
        )
        dcf_model.set_cell(self.ws[1], f"{dcf_static.letters[col]}4", "m")
        dcf_model.set_cell(self.ws[1], f"{dcf_static.letters[col+1]}4", "b")
        rev_pred_type = "growth" if self.growth else "linear"
        self.get_growth("Date", "Revenue", pred_type=rev_pred_type)
        self.get_growth("Revenue", "Cost of Revenue")
        self.get_sum("Gross Profit", "Revenue", [], ["Cost of Revenue"])
        self.get_growth("Revenue", "Selling, General & Admin", True)
        self.get_growth("Revenue", "Research & Development", True)
        self.get_growth("Revenue", "Other Operating Expenses")
        self.get_sum(
            "Operating Income",
            "Gross Profit",
            [],
            [
                "Selling, General & Admin",
                "Research & Development",
                "Other Operating Expenses",
            ],
        )
        self.get_growth("Revenue", "Preferred Dividends")
        self.get_growth("Revenue", "Interest Expense / Income")
        self.get_growth("Revenue", "Other Expense / Income")
        self.get_growth("Operating Income", "Income Tax")
        self.get_sum(
            "Net Income",
            "Operating Income",
            [],
            ["Interest Expense / Income", "Other Expense / Income", "Income Tax"],
        )
        self.custom_exp(
            "Preferred Dividends",
            "Preferred Dividends are not important in a DCF so we do not attempt to predict them.",
        )
        self.get_growth("Revenue", "Cash & Equivalents", True)
        self.get_growth("Revenue", "Short-Term Investments", True)
        self.get_sum(
            "Cash & Cash Equivalents",
            "Cash & Equivalents",
            ["Short-Term Investments"],
            [],
        )
        self.get_growth("Revenue", "Receivables", True)
        self.get_growth("Revenue", "Inventory", True)
        self.get_growth("Revenue", "Other Current Assets")
        self.get_sum(
            "Total Current Assets",
            "Cash & Cash Equivalents",
            ["Receivables", "Inventory", "Other Current Assets"],
            [],
        )
        self.get_growth("Revenue", "Property, Plant & Equipment", True)
        self.get_growth("Revenue", "Long-Term Investments", True)
        self.get_growth("Revenue", "Goodwill and Intangibles", True)
        self.get_growth("Revenue", "Other Long-Term Assets")
        self.get_sum(
            "Total Long-Term Assets",
            "Property, Plant & Equipment",
            [
                "Long-Term Investments",
                "Goodwill and Intangibles",
                "Other Long-Term Assets",
            ],
            [],
        )
        self.get_sum(
            "Total Assets", "Total Current Assets", ["Total Long-Term Assets"], []
        )
        self.get_growth("Revenue", "Accounts Payable")
        self.get_growth("Revenue", "Deferred Revenue")
        self.get_growth("Revenue", "Current Debt")
        self.get_growth("Revenue", "Other Current Liabilities")
        self.get_sum(
            "Total Current Liabilities",
            "Accounts Payable",
            ["Deferred Revenue", "Current Debt", "Other Current Liabilities"],
            [],
        )
        self.get_sum(
            "Long-Term Debt",
            "Total Assets",
            [],
            [
                "Total Current Liabilities",
                "Other Long-Term Liabilities",
                "Shareholders' Equity",
            ],
            text=(
                "This is the plug. For more information on plugs visit https://corporatefina"
                "nceinstitute.com/resources/questions/model-questions/financial-modeling-plug/"
            ),
        )  # This is the plug
        self.get_growth("Revenue", "Other Long-Term Liabilities")
        self.get_sum(
            "Total Long-Term Liabilities",
            "Long-Term Debt",
            ["Other Long-Term Liabilities"],
            [],
        )
        self.get_sum(
            "Total Liabilities",
            "Total Current Liabilities",
            ["Total Long-Term Liabilities"],
            [],
        )
        self.get_growth("Revenue", "Common Stock")
        col = self.info["len_data"] + 1
        rer = self.title_to_row("Retained Earnings")
        nir = self.title_to_row("Net Income")
        for i in range(self.info["len_pred"]):
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+i]}{rer}",
                f"={dcf_static.letters[col+i]}{nir}+{dcf_static.letters[col+i-1]}{rer}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

        self.get_growth("Revenue", "Comprehensive Income")
        self.get_sum(
            "Shareholders' Equity",
            "Common Stock",
            ["Retained Earnings", "Comprehensive Income"],
            [],
        )
        self.get_sum(
            "Total Liabilities and Equity",
            "Total Liabilities",
            ["Shareholders' Equity"],
            [],
        )

        dcf_model.set_cell(
            self.ws[1],
            "A65",
            (
                "Warning: Stock Analysis does not have all of the cash flow items included. Operating"
                ", Financing, and Investing Cash Flows may not add up to total cash flows."
            ),
            font=dcf_static.red,
        )

    @log_start_end(log=logger)
    def create_dcf(self):
        self.ws[2]["A5"] = "Net Income"
        self.ws[2]["A6"] = "Change in NWC"
        self.ws[2]["A7"] = "Change in Capex"
        self.ws[2]["A8"] = "Preferred Dividends"
        self.ws[2]["A9"] = "Free Cash Flows"
        r = 4

        c1 = dcf_static.letters[self.info["len_pred"] + 3]
        c2 = dcf_static.letters[self.info["len_pred"] + 4]
        c3 = dcf_static.letters[self.info["len_pred"] + 5]
        for i in range(self.info["len_pred"]):
            j = 1 + i + self.info["len_data"]
            cols = dcf_static.letters
            dcf_model.set_cell(
                self.ws[2],
                f"{cols[1+i]}4",
                f"=Financials!{cols[j]}4",
                font=dcf_static.bold_font,
            )
            dcf_model.set_cell(
                self.ws[2],
                f"{cols[1+i]}5",
                f"=Financials!{cols[j]}{self.title_to_row('Net Income')}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

            dcf_model.set_cell(
                self.ws[2],
                f"{dcf_static.letters[1+i]}6",
                (
                    f"=Financials!{cols[j]}{self.title_to_row('Total Current Assets')}"
                    f"-Financials!{cols[j-1]}{self.title_to_row('Total Current Assets')}"
                    f"-Financials!{cols[j]}{self.title_to_row('Total Current Liabilities')}"
                    f"+Financials!{cols[j-1]}{self.title_to_row('Total Current Liabilities')}"
                ),
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
            dcf_model.set_cell(
                self.ws[2],
                f"{dcf_static.letters[1+i]}7",
                (
                    f"=Financials!{cols[j]}{self.title_to_row('Total Long-Term Assets')}"
                    f"-Financials!{cols[j-1]}{self.title_to_row('Total Long-Term Assets')}"
                ),
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
            dcf_model.set_cell(
                self.ws[2],
                f"{dcf_static.letters[1+i]}8",
                f"=Financials!{cols[j]}{self.title_to_row('Preferred Dividends')}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
            dcf_model.set_cell(
                self.ws[2],
                f"{cols[1+i]}9",
                f"={cols[1+i]}5-{cols[1+i]}6-{cols[1+i]}7-{cols[1+i]}8",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
                font=dcf_static.bold_font,
                border=dcf_static.thin_border_top,
            )
            dcf_model.set_cell(
                self.ws[2],
                f"{cols[1+self.info['len_pred']]}9",
                f"=({cols[self.info['len_pred']]}9*(1+{c2}"
                f"{r+15}))/({c2}{r+11}-{c2}{r+15})",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

        dcf_model.set_cell(
            self.ws[2],
            f"{c1}{r-2}",
            "Note: We do not allow r values to go below 0.5%.",
            font=dcf_static.red,
        )
        self.ws[2].merge_cells(f"{c1}{r}:{c2}{r}")
        for x in [c1, c2]:
            dcf_model.set_cell(
                self.ws[2], f"{x}{r}", border=dcf_static.thin_border_bottom
            )
        dcf_model.set_cell(
            self.ws[2],
            f"{c1}{r}",
            "Discount Rate",
            alignment=dcf_static.center,
            border=dcf_static.thin_border_bottom,
        )

        # CAPM
        dcf_model.set_cell(self.ws[2], f"{c1}{r+1}", "Risk Free Rate")
        dcf_model.set_cell(
            self.ws[2],
            f"{c2}{r+1}",
            float(self.data["t_bill"]) / 100,
            num_form=FORMAT_PERCENTAGE_00,
        )
        self.custom_exp(r + 1, "Pulled from US Treasurey.", 2, f"{c3}")
        dcf_model.set_cell(self.ws[2], f"{c1}{r+2}", "Market Rate")
        dcf_model.set_cell(
            self.ws[2], f"{c2}{r+2}", 0.08, num_form=FORMAT_PERCENTAGE_00
        )
        self.custom_exp(
            r + 2, "Average return of the S&P 500 is 8% [Investopedia]", 2, f"{c3}"
        )
        dcf_model.set_cell(self.ws[2], f"{c1}{r+3}", "Beta")
        if self.data["info"]["beta"] is None:
            dcf_model.set_cell(self.ws[2], f"{c2}{r+3}", float(1))
            self.custom_exp(
                r + 3, "Warning: Beta not found. Assumed a beta of one.", 2, f"{c3}"
            )
            self.data["info"]["beta"] = 1
        else:
            dcf_model.set_cell(
                self.ws[2], f"{c2}{r+3}", float(self.data["info"]["beta"])
            )
            self.custom_exp(r + 3, "Beta from yahoo finance", 2, f"{c3}")
        dcf_model.set_cell(self.ws[2], f"{c1}{r+4}", "r (CAPM)")
        dcf_model.set_cell(
            self.ws[2],
            f"{c2}{r+4}",
            f"=max((({c2}{r+2}-{c2}{r+1})*{c2}{r+3})+{c2}{r+1},0.005)",
            num_form=FORMAT_PERCENTAGE_00,
            border=dcf_static.thin_border_top,
            font=dcf_static.bold_font,
        )

        # Fama French
        dcf_model.set_cell(self.ws[2], f"{c1}{r+7}", "Fama French")
        dcf_model.set_cell(
            self.ws[2],
            f"{c2}{r+7}",
            f"=max({self.data['r_ff']},0.005)",
            num_form=FORMAT_PERCENTAGE_00,
        )
        self.custom_exp(
            r + 7,
            (
                "Calculated using the Fama and French Three-Factor model. For more"
                "information visit https://www.investopedia.com/terms/f/famaandfrenchthreefactormodel.asp."
            ),
            2,
            f"{c3}",
        )

        # Decide
        for x in [c1, c2]:
            dcf_model.set_cell(
                self.ws[2], f"{x}{r+9}", border=dcf_static.thin_border_bottom
            )
        self.ws[2].merge_cells(f"{c1}{r+9}:{c2}{r+9}")
        dcf_model.set_cell(
            self.ws[2],
            f"{c1}{r+9}",
            "Choose model",
            border=dcf_static.thin_border_bottom,
            alignment=dcf_static.center,
            num_form=FORMAT_PERCENTAGE_00,
        )
        dcf_model.set_cell(self.ws[2], f"{c1}{r+10}", "Model")
        dcf_model.set_cell(self.ws[2], f"{c2}{r+10}", "Fama French")
        dcf_model.set_cell(self.ws[2], f"{c3}{r+10}", "Type 'Fama French' or 'CAPM'")
        dcf_model.set_cell(self.ws[2], f"{c1}{r+11}", "r")
        dcf_model.set_cell(
            self.ws[2],
            f"{c2}{r+11}",
            f'=if({c2}{r+10}="Fama French",{c2}{r+7},if({c2}{r+10}="CAPM",{c2}{r+4},"Invalid Selection"))',
            num_form=FORMAT_PERCENTAGE_00,
        )

        dcf_model.set_cell(self.ws[2], f"{c1}{r+15}", "Long Term Growth")
        dcf_model.set_cell(
            self.ws[2],
            f"{c2}{r+15}",
            f"=min(0.04,{c2}{r+11}*0.9)",
            num_form=FORMAT_PERCENTAGE_00,
        )
        dcf_model.set_cell(self.ws[2], "A11", "Value from Operations")
        dcf_model.set_cell(
            self.ws[2],
            "B11",
            f"=NPV({c2}{r+11},B9:{dcf_static.letters[self.info['len_pred']+1]}9)",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws[2], "A12", "Cash and Cash Equivalents")
        dcf_model.set_cell(
            self.ws[2],
            "B12",
            f"=financials!{dcf_static.letters[self.info['len_data']]}{self.title_to_row('Cash & Cash Equivalents')}",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws[2], "A13", "Intrinsic Value (sum)")
        dcf_model.set_cell(
            self.ws[2],
            "B13",
            "=B11+B12",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws[2], "A14", "Debt Obligations")
        dcf_model.set_cell(
            self.ws[2],
            "B14",
            f"=financials!{dcf_static.letters[self.info['len_data']]}"
            + f"{self.title_to_row('Total Long-Term Liabilities')}",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws[2], "A15", "Firm value without debt")
        dcf_model.set_cell(
            self.ws[2],
            "B15",
            (
                f"=max(B13-B14,"
                f"Financials!{dcf_static.letters[self.info['len_data']]}"
                + f"{self.title_to_row('Total Assets')}"
                f"-Financials!{dcf_static.letters[self.info['len_data']]}"
                + f"{self.title_to_row('Total Liabilities')})"
            ),
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(
            self.ws[2],
            "C15",
            *(
                f"=if((B13-B14)>"
                f"(Financials!{dcf_static.letters[self.info['len_data']]}"
                + f"{self.title_to_row('Total Assets')}"
                f"-Financials!{dcf_static.letters[self.info['len_data']]}"
                + f"{self.title_to_row('Total Liabilities')}),"
                '"","Note: Total assets minus total liabilities exceeds projected firm value without debt.'
                + ' Value shown is total assets minus total liabilities.")',
            ),
            font=dcf_static.red,
        )
        dcf_model.set_cell(self.ws[2], "A16", "Shares Outstanding")
        dcf_model.set_cell(self.ws[2], "B16", int(self.data["f_info"]["shares"]))
        dcf_model.set_cell(self.ws[2], "A17", "Shares Price")
        dcf_model.set_cell(
            self.ws[2],
            "B17",
            f"=(B15*{self.info['rounding']})/B16",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws[2], "A18", "Actual Price")
        dcf_model.set_cell(self.ws[2], "B18", float(self.data["f_info"]["last_price"]))

    @log_start_end(log=logger)
    def create_header(self, ws: Workbook):
        for i in range(10):
            dcf_model.set_cell(
                ws, f"{dcf_static.letters[i]}1", border=dcf_static.thin_border
            )

        ws.merge_cells("A1:J1")
        dcf_model.set_cell(
            ws,
            "A1",
            f"OpenBB Terminal Analysis: {self.info['symbol'].upper()}",
            font=Font(color="04cca8", size=20),
            border=dcf_static.thin_border,
            alignment=dcf_static.center,
        )
        dcf_model.set_cell(
            ws, "A2", f"DCF for {self.info['symbol']} generated on {self.data['now']}"
        )

    @log_start_end(log=logger)
    def run_audit(self):
        start = 67
        for i, value in enumerate(dcf_static.sum_rows):
            dcf_model.set_cell(self.ws[1], f"A{start + i}", value)

        self.ws[1].merge_cells(f"A{start-2}:K{start-2}")
        dcf_model.set_cell(
            self.ws[1],
            f"A{start - 2}",
            "Financial Statement Audit",
            font=Font(color="FF0000"),
            alignment=dcf_static.center,
        )
        dcf_model.set_cell(
            self.ws[1],
            f"A{start - 1}",
            "Audit ensures data integrity. Numbers should be 0 (with slight rounding difference).",
        )

        self.get_sum(start, "Revenue", [], ["Cost of Revenue", "Gross Profit"], True)
        self.get_sum(
            start + 1,
            "Gross Profit",
            [],
            [
                "Selling, General & Admin",
                "Research & Development",
                "Other Operating Expenses",
                "Operating Income",
            ],
            True,
        )
        self.get_sum(
            start + 2,
            "Operating Income",
            [],
            [
                "Interest Expense / Income",
                "Other Expense / Income",
                "Income Tax",
                "Net Income",
            ],
            True,
        )
        self.get_sum(
            start + 3,
            "Cash & Equivalents",
            ["Short-Term Investments"],
            ["Cash & Cash Equivalents"],
            True,
        )
        self.get_sum(
            start + 4,
            "Cash & Cash Equivalents",
            ["Receivables", "Inventory", "Other Current Assets"],
            ["Total Current Assets"],
            True,
        )
        self.get_sum(
            start + 5,
            "Property, Plant & Equipment",
            [
                "Long-Term Investments",
                "Goodwill and Intangibles",
                "Other Long-Term Assets",
            ],
            ["Total Long-Term Assets"],
            True,
        )
        self.get_sum(
            start + 6,
            "Total Current Assets",
            ["Total Long-Term Assets"],
            ["Total Assets"],
            True,
        )
        self.get_sum(
            start + 7,
            "Accounts Payable",
            ["Deferred Revenue", "Current Debt", "Other Current Liabilities"],
            ["Total Current Liabilities"],
            True,
        )
        self.get_sum(
            start + 8,
            "Long-Term Debt",
            ["Other Long-Term Liabilities"],
            ["Total Long-Term Liabilities"],
            True,
        )
        self.get_sum(
            start + 9,
            "Total Current Liabilities",
            ["Total Long-Term Liabilities"],
            ["Total Liabilities"],
            True,
        )
        self.get_sum(
            start + 10,
            "Common Stock",
            ["Retained Earnings", "Comprehensive Income"],
            ["Shareholders' Equity"],
            True,
        )
        self.get_sum(
            start + 11,
            "Total Liabilities",
            ["Shareholders' Equity"],
            ["Total Liabilities and Equity"],
            True,
        )
        self.get_sum(
            start + 12,
            "Net Income",
            [
                "Depreciation & Amortization",
                "Share-Based Compensation",
                "Other Operating Activities",
            ],
            ["Operating Cash Flow"],
            True,
        )
        self.get_sum(
            start + 13,
            "Capital Expenditures",
            ["Acquisitions", "Change in Investments", "Other Investing Activities"],
            ["Investing Cash Flow"],
            True,
        )
        self.get_sum(
            start + 14,
            "Dividends Paid",
            [
                "Share Issuance / Repurchase",
                "Debt Issued / Paid",
                "Other Financing Activities",
            ],
            ["Financing Cash Flow"],
            True,
        )
        self.get_sum(
            start + 15,
            "Operating Cash Flow",
            ["Investing Cash Flow", "Financing Cash Flow"],
            ["Net Cash Flow"],
            True,
        )

    @log_start_end(log=logger)
    def get_growth(
        self, x_ind: str, y_ind: str, no_neg: bool = False, pred_type: str = "linear"
    ):
        """Add growth to a column. Usually this is linear but other options are allowed

        x_ind: str
            The x variable to use. This is unused if growth is the pred_type
        y_ind: str
            The y variable to use
        no_neg: bool
            Whether or not the value can have negative numbers (profit can be revenue cannot)
        pred_type: str
            This is assumed to be linear but growth is allowed as well
        """

        x_type = "IS" if x_ind in self.df["IS"].index else "BS"
        y_type = "IS" if y_ind in self.df["IS"].index else "BS"
        x_df = self.df["IS"] if x_type == "IS" else self.df["BS"]
        y_df = self.df["IS"] if y_type == "IS" else self.df["BS"]

        analy_message = ""
        row1 = (
            y_df.index.get_loc(y_ind)
            + 1
            + (self.starts["IS"] if y_type == "IS" else self.starts["BS"])
        )

        col = self.info["len_pred"] + self.info["len_data"] + 3
        if pred_type == "linear":
            pre_x = (
                x_df.columns.to_numpy()
                if x_ind == "Date"
                else x_df.loc[x_ind].to_numpy()
            )

            vfunc = np.vectorize(dcf_model.string_float)
            pre_x = vfunc(pre_x)

            if x_ind == "Date":
                pre_x = pre_x - np.min(pre_x)
            x = pre_x.reshape((-1, 1))
            pre_y = y_df.loc[y_ind].to_numpy()
            y = vfunc(pre_y)
            model = LinearRegression().fit(x, y)
            r_sq = model.score(x, y)
            r = abs(r_sq**0.5)

            if r > 0.9:
                strength = "very strong"
            elif r > 0.7:
                strength = "strong"
            elif r > 0.5:
                strength = "moderate"
            elif r > 0.3:
                strength = "weak"
            else:
                strength = "very weak"
            analy_message = (
                f"The correlation between {x_ind.lower()} and {y_ind.lower()}"
                f" is {strength} with a correlation coefficient of {r:.4f}."
            )
            dcf_model.set_cell(
                self.ws[1], f"{dcf_static.letters[col]}{row1}", float(model.coef_)
            )
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+1]}{row1}",
                float(model.intercept_),
            )
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+2]}{row1}",
                dcf_static.letters[self.letter],
                font=dcf_static.red,
            )
            dcf_model.set_cell(
                self.ws[3],
                f"A{self.letter+4}",
                dcf_static.letters[self.letter],
                font=dcf_static.red,
            )
        elif pred_type == "growth":
            pre_y = y_df.loc[y_ind].to_numpy()
            fin_y = []
            for item in pre_y:
                try:
                    fin_y.append(float(item))
                except TypeError:
                    fin_y.append(0)
            overall_growth = (fin_y[-1] - fin_y[0]) / fin_y[0]
            yearly_growth = overall_growth / len(fin_y)
            analy_message = (
                f"Assuming growth percentage of {(yearly_growth * 100):.2f}%"
            )
            dcf_model.set_cell(
                self.ws[1], f"{dcf_static.letters[col]}{row1}", yearly_growth
            )
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+2]}{row1}",
                dcf_static.letters[self.letter],
                font=dcf_static.red,
            )

        dcf_model.set_cell(self.ws[3], f"B{self.letter+4}", analy_message)

        col = self.info["len_data"] + 1
        for i in range(self.info["len_pred"]):
            if pred_type == "growth":
                row_n = (
                    y_df.index.get_loc(y_ind) + 1 + self.starts["IS"]
                    if y_type == "IS"
                    else self.starts["BS"]
                )
                base = (
                    f"({dcf_static.letters[col+i-1]}{row1}* (1+"
                    f"{dcf_static.letters[col+self.info['len_pred']+2]}{row1}))"
                )

            elif x_ind == "Date":
                base = (
                    f"(({dcf_static.letters[col+i]}4-B4)*"
                    f"{dcf_static.letters[col+self.info['len_pred']+2]}"
                    f"{row1})+{dcf_static.letters[col+self.info['len_pred']+3]}{row1}"
                )
            else:
                row_n = (
                    x_df.index.get_loc(x_ind) + 1 + self.starts["IS"]
                    if x_type == "IS"
                    else self.starts["BS"]
                )
                base = (
                    f"({dcf_static.letters[col+i]}{row_n}*"
                    f"{dcf_static.letters[col+self.info['len_pred']+2]}{row1})"
                    f"+{dcf_static.letters[col+self.info['len_pred']+3]}{row1}"
                )
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+i]}{row1}",
                f"=max({base},0)" if no_neg else f"={base}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

        self.letter += 1

    @log_start_end(log=logger)
    def get_sum(
        self,
        row: Union[int, str],
        first: str,
        adds: List[str],
        subtracts: List[str],
        audit: bool = False,
        text: Optional[str] = None,
    ):
        col = 1 if audit else self.info["len_data"] + 1
        for i in range(self.info["len_data"] if audit else self.info["len_pred"]):
            sum_formula = f"={dcf_static.letters[col+i]}{self.title_to_row(first)}"
            for item in adds:
                sum_formula += f"+{dcf_static.letters[col+i]}{self.title_to_row(item)}"
            for item in subtracts:
                sum_formula += f"-{dcf_static.letters[col+i]}{self.title_to_row(item)}"
            rowI = row if isinstance(row, int) else self.title_to_row(row)
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+i]}{rowI}",
                sum_formula,
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
        if text:
            self.custom_exp(row, text)

    def title_to_row(self, title: str) -> int:
        df = (
            self.df["IS"]
            if title in self.df["IS"].index
            else self.df["BS"]
            if title in self.df["BS"].index
            else self.df["CF"]
        )
        ind = (
            df.index.get_loc(title)
            + 1
            + (
                self.starts["IS"]
                if title in self.df["IS"].index
                else self.starts["BS"]
                if title in self.df["BS"].index
                else self.starts["CF"]
            )
        )
        return ind

    @log_start_end(log=logger)
    def custom_exp(
        self, row: Union[int, str], text: str, ws: int = 1, column: Optional[str] = None
    ):
        if ws == 1:
            rowT = row if isinstance(row, int) else self.title_to_row(row)
            col = self.info["len_pred"] + self.info["len_data"] + 3
            dcf_model.set_cell(
                self.ws[1],
                f"{dcf_static.letters[col+2]}{rowT}",
                dcf_static.letters[self.letter],
                font=dcf_static.red,
            )
        if ws == 2:
            dcf_model.set_cell(
                self.ws[2],
                f"{column}{row}",
                dcf_static.letters[self.letter],
                font=dcf_static.red,
            )

        dcf_model.set_cell(
            self.ws[3],
            f"A{self.letter+4}",
            dcf_static.letters[self.letter],
            font=dcf_static.red,
        )
        dcf_model.set_cell(self.ws[3], f"B{self.letter+4}", text)
        self.letter += 1

    @log_start_end(log=logger)
    def add_ratios(self):
        self.ws[4] = self.wb.create_sheet("Ratios")
        self.create_header(self.ws[4])
        self.ws[4].column_dimensions["A"].width = 27
        dcf_model.set_cell(self.ws[4], "B4", "Sector:")
        dcf_model.set_cell(self.ws[4], "C4", self.data["info"]["sector"])
        similar_data = dcf_model.get_similar_dfs(
            self.info["symbol"],
            self.data["info"],
            self.info["max_similars"],
            self.info["no_filter"],
        )
        similar_data.insert(
            0, [self.info["symbol"], [self.df["BS"], self.df["IS"], self.df["CF"]]]
        )
        row = 6
        for val in similar_data:
            self.ws[4].merge_cells(f"A{row}:J{row}")

            # Header columns formatted as row index, text, format type
            titles = [
                [0, val[0], 1],
                [1, "Liquidity Ratios", 2],
                [2, "Current Ratio", 0],
                [3, "Quick Ratio", 0],
                [5, "Activity Ratios", 2],
                [6, "AR Turnover", 0],
                [7, "Days Sales in AR", 0],
                [8, "Inventory Turnover", 0],
                [9, "Days in Inventory", 0],
                [10, "Average Payable Turnover", 0],
                [11, "Days of Payables Outstanding", 0],
                [12, "Cash Conversion Cycle", 0],
                [13, "Asset Turnover", 0],
                [15, "Profitability Ratios", 2],
                [16, "Profit Margin", 0],
                [17, "Return on Assets", 0],
                [18, "Return on Equity", 0],
                [19, "Return on Sales", 0],
                [20, "Gross Margin", 0],
                [21, "Operating Cash Flow Ratio", 0],
                [23, "Coverage Ratios", 2],
                [24, "Debt-to-Equity", 0],
                [25, "Total Debt Ratio", 0],
                [26, "Equity Multiplier", 0],
                [27, "Times Interest Earned", 0],
                [29, "Investor Ratios", 2],
                [30, "Earnings Per Share", 0],
                [31, "Price Earnings Ratio", 0],
            ]

            for j in titles:
                if j[2] == 1:
                    align = dcf_static.center
                    dcf_model.set_cell(
                        self.ws[4], f"A{row+j[0]}", j[1], alignment=align
                    )
                elif j[2] == 2:
                    border = dcf_static.thin_border
                    font = dcf_static.bold_font
                    dcf_model.set_cell(
                        self.ws[4], f"A{row+j[0]}", j[1], border=border, font=font
                    )
                else:
                    dcf_model.set_cell(self.ws[4], f"A{row+j[0]}", j[1])

            for j in range(val[1][0].shape[1] - 1):
                lt = dcf_static.letters[j + 1]

                cace1 = dcf_model.get_value(val[1][0], "Cash & Cash Equivalents", j)[1]
                ar0, ar1 = dcf_model.get_value(val[1][0], "Receivables", j)
                inv0, inv1 = dcf_model.get_value(val[1][0], "Inventory", j)
                ca1 = dcf_model.get_value(val[1][0], "Total Current Assets", j)[1]
                ta0, ta1 = dcf_model.get_value(val[1][0], "Total Assets", j)
                ap0, ap1 = dcf_model.get_value(val[1][0], "Accounts Payable", j)
                cl1 = dcf_model.get_value(val[1][0], "Total Current Liabilities", j)[1]
                tl1 = dcf_model.get_value(val[1][0], "Total Liabilities", j)[1]
                te0, te1 = dcf_model.get_value(val[1][0], "Shareholders' Equity", j)
                sls1 = dcf_model.get_value(val[1][1], "Revenue", j)[1]
                cogs1 = dcf_model.get_value(val[1][1], "Cost of Revenue", j)[1]
                inte1 = dcf_model.get_value(val[1][1], "Interest Expense / Income", j)[
                    1
                ]
                tax1 = dcf_model.get_value(val[1][1], "Income Tax", j)[1]
                ni1 = dcf_model.get_value(val[1][1], "Net Income", j)[1]
                pdiv1 = dcf_model.get_value(val[1][1], "Preferred Dividends", j)[1]
                opcf1 = dcf_model.get_value(val[1][2], "Operating Cash Flow", j)[1]

                info, outstand = self.data["f_info"], float(
                    self.data["f_info"]["shares"]
                )

                # Enter row offset, number to display, and format number
                rows = [
                    [1, int(val[1][0].columns[j + 1]), 1],
                    [2, dcf_model.frac(ca1, cl1), 0],
                    [3, dcf_model.frac(cace1 + ar1, cl1), 0],
                    [6, dcf_model.frac(sls1, (ar0 + ar1) / 2), 0],
                    [7, dcf_model.frac(ar1, sls1 / 365), 0],
                    [8, dcf_model.frac(cogs1, (inv0 + inv1) / 2), 0],
                    [9, dcf_model.frac(inv1, cogs1 / 365), 0],
                    [10, dcf_model.frac(cogs1, (ap0 + ap1) / 2), 0],
                    [11, dcf_model.frac(ap1, cogs1 / 365), 0],
                    [
                        12,
                        "N/A"
                        if sls1 == 0 or cogs1 == 0
                        else dcf_model.frac(ar1, sls1 / 365)
                        + dcf_model.frac(inv1, cogs1 / 365)
                        - dcf_model.frac(ap1, cogs1 / 365),
                        0,
                    ],
                    [13, dcf_model.frac(sls1, (ta0 + ta1) / 2), 0],
                    [16, dcf_model.frac(ni1, sls1), 0],
                    [17, dcf_model.frac(ni1, (ar0 + ar1) / 2), 0],
                    [18, dcf_model.frac(ni1, (te0 + te1) / 2), 0],
                    [19, dcf_model.frac(ni1 + inte1 + tax1, sls1), 0],
                    [20, dcf_model.frac(sls1 - cogs1, sls1), 0],
                    [21, dcf_model.frac(opcf1, cl1), 0],
                    [24, dcf_model.frac(tl1, te1), 0],
                    [25, dcf_model.frac(tl1, ta1), 0],
                    [26, dcf_model.frac(ta1, te1), 0],
                    [27, dcf_model.frac(ni1 + inte1 + tax1, inte1), 0],
                    [
                        30,
                        dcf_model.frac((ni1 - pdiv1) * self.info["rounding"], outstand),
                        0,
                    ],
                    [
                        31,
                        dcf_model.frac(
                            float(info["regular_market_previous_close"]) * outstand,
                            (ni1 - pdiv1) * self.info["rounding"],
                        ),
                        0,
                    ],
                ]

                for k in rows:
                    if k[2] == 1:
                        font = dcf_static.bold_font
                        dcf_model.set_cell(
                            self.ws[4], f"{lt}{row+k[0]}", k[1], font=font
                        )
                    else:
                        dcf_model.set_cell(self.ws[4], f"{lt}{row+k[0]}", k[1])

            row += 35
