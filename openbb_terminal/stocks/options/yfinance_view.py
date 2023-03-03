"""Yfinance options view"""
__docformat__ = "numpy"

import logging
import os
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Union

import numpy as np
import pandas as pd
from openpyxl import Workbook
from scipy.spatial import Delaunay
from scipy.stats import binom

from openbb_terminal import OpenBBFigure
from openbb_terminal.core.config.paths import MISCELLANEOUS_DIRECTORY
from openbb_terminal.core.plots.config.openbb_styles import (
    PLT_3DMESH_COLORSCALE,
    PLT_3DMESH_HOVERLABEL,
    PLT_3DMESH_SCENE,
)
from openbb_terminal.decorators import log_start_end
from openbb_terminal.helper_funcs import (
    excel_columns,
    export_data,
    get_rf,
    print_rich_table,
)
from openbb_terminal.rich_config import console
from openbb_terminal.stocks.options import op_helpers, yfinance_model
from openbb_terminal.stocks.options.yfinance_model import (
    generate_data,
    get_option_chain,
    get_price,
)

# pylint: disable=C0302, R0913, too-many-arguments


logger = logging.getLogger(__name__)


def header_fmt(header: str) -> str:
    """
    Formats strings to appear as titles

    Parameters
    ----------
    header: str
        The string to be formatted

    Returns
    -------
    new_header: str
        The clean string to use as a header
    """

    words = re.findall("[A-Z][^A-Z]*", header)
    if words == []:
        words = [header]
    new_header = " ".join(words)
    new_header = new_header.replace("_", " ")
    return new_header.title()


@log_start_end(log=logger)
def plot_plot(
    symbol: str,
    expiry: str,
    put: bool = False,
    x: str = "s",
    y: str = "iv",
    custom: str = "",
    export: str = "",
    sheet_name: Optional[str] = None,
    external_axes: bool = False,
) -> Union[OpenBBFigure, None]:
    """Generate a graph custom graph based on user input

    Parameters
    ----------
    symbol: str
        Stock ticker symbol
    expiry: str
        Option expiration
    x: str
        variable to display in x axis, choose from:
        ltd, s, lp, b, a, c, pc, v, oi, iv
    y: str
        variable to display in y axis, choose from:
        ltd, s, lp, b, a, c, pc, v, oi, iv
    custom: str
        type of plot
    put: bool
        put option instead of call
    sheet_name: str
        Optionally specify the name of the sheet the data is exported to.
    export: str
        type of data to export
    external_axes : bool, optional
        Whether to return the figure object or not, by default False
    """
    convert = {
        "ltd": "lastTradeDate",
        "s": "strike",
        "lp": "lastPrice",
        "b": "bid",
        "a": "ask",
        "c": "change",
        "pc": "percentChange",
        "v": "volume",
        "oi": "openInterest",
        "iv": "impliedVolatility",
    }
    if custom == "smile":
        x = "strike"
        y = "impliedVolatility"
    else:
        if x is None:
            return console.print("[red]Invalid option sent for x-axis[/red]\n")
        if y is None:
            return console.print("[red]Invalid option sent for y-axis[/red]\n")
        if x in convert:
            x = convert[x]
        else:
            x = "strike"
            console.print(
                f"[red]'{x}' is not a valid option. Defaulting to `strike`.[/red]\n"
            )
        if y in convert:
            y = convert[y]
        else:
            y = "impliedVolatility"
            console.print(
                f"[red]'{y}' is not a valid option. Defaulting to `impliedVolatility`.[/red]\n"
            )

    varis = op_helpers.opt_chain_cols
    chain = yfinance_model.get_option_chain(symbol, expiry)
    values = chain.puts if put else chain.calls

    x_data = values[x]
    y_data = values[y]

    option = "puts" if put else "calls"
    fig = OpenBBFigure(
        title=f"{varis[y]['label']} vs. {varis[x]['label']} for {symbol} {option} on {expiry}",
        xaxis_title=varis[x]["label"],
        yaxis_title=varis[y]["label"],
    )
    fig.add_scatter(x=x_data, y=y_data, mode="lines+markers")

    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "plot",
        sheet_name=sheet_name,
        figure=fig,
    )

    return fig.show(external=external_axes)


@log_start_end(log=logger)
def plot_payoff(
    current_price: float,
    options: List[Dict[Any, Any]],
    underlying: float,
    symbol: str,
    expiry: str,
    external_axes: bool = False,
) -> Union[OpenBBFigure, None]:
    """Generate a graph showing the option payoff diagram"""
    x, yb, ya = generate_data(current_price, options, underlying)

    plot = OpenBBFigure(
        title=f"Option Payoff Diagram for {symbol} on {expiry}",
        xaxis=dict(tickformat="${x:.2f}", title="Underlying Asset Price at Expiration"),
        yaxis=dict(tickformat="${x:.2f}", title="Profit"),
    )
    plot.add_scatter(x=x, y=yb, name="Payoff Before Premium")
    if ya:
        plot.add_scatter(x=x, y=ya, name="Payoff After Premium")
    else:
        plot.data[0].name = "Payoff"

    return plot.show(external=external_axes)


@log_start_end(log=logger)
def show_parity(
    symbol: str,
    expiry: str,
    put: bool = False,
    ask: bool = False,
    mini: Optional[float] = None,
    maxi: Optional[float] = None,
    export: str = "",
    sheet_name: Optional[str] = None,
) -> None:
    """Prints options and whether they are under or over priced [Source: Yahoo Finance]

    Parameters
    ----------
    symbol : str
        Ticker symbol to get expirations for
    expiration : str
        Expiration to use for options
    put : bool
        Whether to use puts or calls
    ask : bool
        Whether to use ask or lastPrice
    mini : float
        Minimum strike price to show
    maxi : float
        Maximum strike price to show
    export : str
        Export data
    """
    r_date = datetime.strptime(expiry, "%Y-%m-%d").date()
    delta = (r_date - date.today()).days
    rate = ((1 + get_rf()) ** (delta / 365)) - 1
    stock = get_price(symbol)

    div_info = yfinance_model.get_dividend(symbol)
    div_dts = div_info.index.values.tolist()

    if div_dts:
        last_div = pd.to_datetime(div_dts[-1])

        avg_div = (
            np.mean(div_info.to_list()[-4:])
            if len(div_dts) > 3
            else np.mean(div_info.to_list())
        )

        next_div = last_div + timedelta(days=91)
        dividends = []
        while next_div < datetime.strptime(expiry, "%Y-%m-%d"):
            day_dif = (next_div - datetime.now()).days
            dividends.append((avg_div, day_dif))
            next_div += timedelta(days=91)
        div_pvs = [x[0] / ((1 + get_rf()) ** (x[1] / 365)) for x in dividends]
        pv_dividend = sum(div_pvs)
    else:
        pv_dividend = 0

    chain = get_option_chain(symbol, expiry)
    name = "ask" if ask else "lastPrice"
    o_type = "put" if put else "call"

    calls = chain.calls[["strike", name]].copy()
    calls = calls.rename(columns={name: "callPrice"})
    puts = chain.puts[["strike", name]].copy()
    puts = puts.rename(columns={name: "putPrice"})

    opts = pd.merge(calls, puts, on="strike")
    opts = opts.dropna()
    opts = opts.loc[opts["callPrice"] * opts["putPrice"] != 0]

    opts["callParity"] = (
        opts["putPrice"] + stock - (opts["strike"] / (1 + rate)) - pv_dividend
    )
    opts["putParity"] = (
        (opts["strike"] / (1 + rate)) + opts["callPrice"] - stock + pv_dividend
    )

    diff = o_type + " Difference"
    opts[diff] = opts[o_type + "Price"] - opts[o_type + "Parity"]
    opts["distance"] = abs(stock - opts["strike"])
    filtered = opts.copy()

    if mini is None:
        mini = filtered.strike.quantile(0.25)
    if maxi is None:
        maxi = filtered.strike.quantile(0.75)

    filtered = filtered.loc[filtered["strike"] >= mini]
    filtered = filtered.loc[filtered["strike"] <= maxi]

    show = filtered[["strike", diff]].copy()

    if ask:
        console.print("Warning: Options with no current ask price not shown.\n")

    print_rich_table(
        show,
        headers=[x.title() for x in show.columns],
        show_index=False,
        title=f"{symbol} Parity",
        export=bool(export),
    )
    console.print(
        "[yellow]Warning: Low volume options may be difficult to trade.[/yellow]"
    )

    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "parity",
        show,
        sheet_name,
    )


@log_start_end(log=logger)
def risk_neutral_vals(
    symbol: str,
    expiry: str,
    data: pd.DataFrame,
    put: bool = False,
    mini: Optional[float] = None,
    maxi: Optional[float] = None,
    risk: Optional[float] = None,
) -> None:
    """Prints current options prices and risk neutral values [Source: Yahoo Finance]

    Parameters
    ----------
    symbol: str
        Ticker symbol to get expirations for
    expiry: str
        Expiration to use for options
    put: bool
        Whether to use puts or calls
    data: pd.DataFrame
        Estimates for stocks prices and probabilities
    mini: float
        Minimum strike price to show
    maxi: float
        Maximum strike price to show
    risk: float
        The risk-free rate for the asset
    """
    chain = (
        get_option_chain(symbol, expiry).puts
        if put
        else get_option_chain(symbol, expiry).calls
    )

    r_date = datetime.strptime(expiry, "%Y-%m-%d").date()
    delta = (r_date - date.today()).days
    vals = []
    if risk is None:
        risk = get_rf()
    for _, row in chain.iterrows():
        vals.append(
            [
                row["strike"],
                row["lastPrice"],
                op_helpers.rn_payoff(row["strike"], data, put, delta, risk),
            ]
        )
    new_df = pd.DataFrame(vals, columns=["Strike", "Last Price", "Value"], dtype=float)
    new_df["Difference"] = new_df["Last Price"] - new_df["Value"]

    if mini is None:
        mini = new_df.Strike.quantile(0.25)
    if maxi is None:
        maxi = new_df.Strike.quantile(0.75)

    new_df = new_df[new_df["Strike"] >= mini]
    new_df = new_df[new_df["Strike"] <= maxi]

    print_rich_table(
        new_df,
        headers=[x.title() for x in new_df.columns],
        show_index=False,
        title="Risk Neutral Values",
    )


@log_start_end(log=logger)
def plot_expected_prices(
    und_vals: List[List[float]],
    p: float,
    symbol: str,
    expiry: str,
    external_axes: bool = False,
) -> Union[OpenBBFigure, None]:
    """Plot expected prices of the underlying asset at expiration

    Parameters
    ----------
    und_vals : List[List[float]]
        The expected underlying values at the expiration date
    p : float
        The probability of the stock price moving upward each round
    symbol : str
        The ticker symbol of the option's underlying asset
    expiry : str
        The expiration for the option
    external_axes : bool, optional
        Whether to return the figure object or not, by default False
    """
    up_moves = list(range(len(und_vals[-1])))
    up_moves.reverse()
    probs = [100 * binom.pmf(r, len(up_moves), p) for r in up_moves]

    fig = OpenBBFigure()

    fig.add_scatter(
        x=und_vals[-1],
        y=probs,
        mode="lines",
        name="Probabilities",
    )

    fig.update_layout(
        title=f"Probabilities for ending prices of {symbol} on {expiry}",
        xaxis_title="Price",
        yaxis_title="Probability",
        xaxis_tickformat="$,.2f",
        yaxis_tickformat=".0%",
    )

    return fig.show(external=external_axes)


@log_start_end(log=logger)
def export_binomial_calcs(
    up: float,
    prob_up: float,
    discount: float,
    und_vals: List[List[float]],
    opt_vals: List[List[float]],
    days: int,
    symbol: str,
) -> None:
    """Create an excel spreadsheet with binomial tables for underlying asset value and option value

    Parameters
    ----------
    up : float
        The stock's increase on an upward move
    prob_up : float
        The probability of an upward move
    discount : float
        The daily discount rate
    und_vals : List[List[float]]
        The underlying asset values at each step
    opt_vals : List[List[float]]
        The values for the option at each step
    days : int
        The number of days until the option expires
    symbol : str
        The ticker symbol for the company
    """

    opts = excel_columns()
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Up Move"
    ws["B1"] = up
    ws["A2"] = "Down Move"
    ws["B2"] = 1 / up
    ws["D1"] = "Prob Up"
    ws["E1"] = prob_up
    ws["D2"] = "Prob Down"
    ws["E2"] = 1 - prob_up
    ws["D3"] = "Discount"
    ws["E3"] = discount
    ws["A4"] = "Binomial Tree for Underlying Values"
    for i, _ in enumerate(und_vals):
        for j, _ in enumerate(und_vals[i]):
            ws[f"{opts[i]}{j+5}"] = und_vals[i][j]

    ws[f"A{days+7}"] = "Binomial Tree for Option Values"
    for i, _ in enumerate(opt_vals):
        for j, _ in enumerate(opt_vals[i]):
            ws[f"{opts[i]}{j+8+days}"] = opt_vals[i][j]

    trypath = str(
        MISCELLANEOUS_DIRECTORY
        / "exports"
        / "stocks"
        / "options"
        / f"{symbol} {datetime.now()}.xlsx"
    )
    wb.save(trypath)
    console.print(
        f"Analysis ran for {symbol}\nPlease look in {trypath} for the file.\n"
    )


@log_start_end(log=logger)
def show_binom(
    symbol: str,
    expiry: str,
    strike: float = 0,
    put: bool = False,
    europe: bool = False,
    export: str = "",
    plot: bool = False,
    vol: Optional[float] = None,
) -> None:
    """Get binomial pricing for option

    Parameters
    ----------
    symbol : str
        The ticker symbol of the option's underlying asset
    expiry : str
        The expiration for the option
    strike : float
        The strike price for the option
    put : bool
        Value a put instead of a call
    europe : bool
        Value a European option instead of an American option
    export : str
        Export the options data to an excel spreadsheet
    plot : bool
        Show a graph of expected ending prices
    vol : float
        The annualized volatility for the underlying asset
    """
    fig = OpenBBFigure()

    up, prob_up, discount, und_vals, opt_vals, days = yfinance_model.get_binom(
        symbol, expiry, strike, put, europe, vol
    )

    if plot or fig.is_image_export(export):
        fig = plot_expected_prices(und_vals, prob_up, symbol, expiry, True)
        export_data(
            export,
            os.path.dirname(os.path.abspath(__file__)),
            "binomial",
            figure=fig,
        )

    if export:
        export_binomial_calcs(up, prob_up, discount, und_vals, opt_vals, days, symbol)

    option = "put" if put else "call"
    console.print(
        f"{symbol} {option} at ${strike:.2f} expiring on {expiry} is worth ${opt_vals[0][0]:.2f}\n"
    )


@log_start_end(log=logger)
def display_vol_surface(
    symbol: str,
    export: str = "",
    sheet_name: Optional[str] = None,
    z: str = "IV",
    external_axes: bool = False,
) -> Union[OpenBBFigure, None]:
    """Display vol surface

    Parameters
    ----------
    symbol : str
        Ticker symbol to get surface for
    export : str
        Format to export data
    z : str
        The variable for the Z axis
    external_axes : bool, optional
        Whether to return the figure object or not, by default False
    """
    data = yfinance_model.get_iv_surface(symbol)
    z = z.upper()

    if data.empty:
        return console.print(f"No options data found for {symbol}.\n")

    z_dict = {
        "IV": (data.impliedVolatility, "Volatility"),
        "OI": (data.openInterest, "Open Interest"),
        "LP": (data.lastPrice, "Last Price"),
    }
    label = z_dict[z][1]

    if z not in z_dict:
        return console.print(
            f"Invalid z value: {z}.\n Valid values are: IV, OI, LP. (case insensitive)\n"
        )

    X = data.dte
    Y = data.strike
    Z = z_dict[z][0]

    points3D = np.vstack((X, Y, Z)).T
    points2D = points3D[:, :2]
    tri = Delaunay(points2D)
    II, J, K = tri.simplices.T

    fig = OpenBBFigure()
    fig.set_title(f"{label} Surface for {symbol.upper()}")

    fig_kwargs = dict(z=Z, x=X, y=Y, i=II, j=J, k=K, intensity=Z)
    fig.add_mesh3d(
        **fig_kwargs,
        colorscale=PLT_3DMESH_COLORSCALE,
        hovertemplate="<b>DTE</b>: %{y} <br><b>Strike</b>: %{x} <br><b>"
        + z_dict[z][1]
        + "</b>: %{z}<extra></extra>",
        showscale=False,
        flatshading=True,
        lighting=dict(
            ambient=0.5, diffuse=0.5, roughness=0.5, specular=0.4, fresnel=0.4
        ),
    )

    tick_kwargs = dict(tickfont=dict(size=13), titlefont=dict(size=16))
    fig.update_layout(
        scene=dict(
            xaxis=dict(title="Strike", **tick_kwargs),
            yaxis=dict(title="DTE", **tick_kwargs),
            zaxis=dict(title=z, **tick_kwargs),
        )
    )
    fig.update_layout(
        margin=dict(l=5, r=10, t=40, b=20),
        title_x=0.5,
        hoverlabel=PLT_3DMESH_HOVERLABEL,
        scene_camera=dict(
            up=dict(x=0, y=0, z=2),
            center=dict(x=0, y=0, z=-0.3),
            eye=dict(x=1.25, y=1.25, z=0.69),
        ),
        scene=PLT_3DMESH_SCENE,
    )

    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "vsurf",
        data,
        sheet_name,
        fig,
        margin=False,
    )

    return fig.show(external=external_axes, margin=False)


@log_start_end(log=logger)
def show_greeks(
    symbol: str,
    expiry: str,
    div_cont: float = 0,
    rf: Optional[float] = None,
    opt_type: int = 1,
    mini: float = -1,
    maxi: float = -1,
    show_all: bool = False,
) -> None:
    """
    Shows the greeks for a given option

    Parameters
    ----------
    symbol: str
        The ticker symbol value of the option
    div_cont: float
        The dividend continuous rate
    expiry: str
        The date of expiration, format "YYYY-MM-DD", i.e. 2010-12-31.
    rf: float
        The risk-free rate
    opt_type: Union[1, -1]
        The option type 1 is for call and -1 is for put
    mini: float
        The minimum strike price to include in the table
    maxi: float
        The maximum strike price to include in the table
    show_all: bool
        Whether to show all greeks
    """

    current_price = get_price(symbol)
    chain = get_option_chain(symbol, expiry)

    min_strike, max_strike = op_helpers.get_strikes(
        min_sp=mini, max_sp=maxi, current_price=current_price
    )

    for option in ["calls", "puts"]:
        attr = getattr(chain, option)
        attr = attr[attr["strike"] >= min_strike]
        attr = attr[attr["strike"] <= max_strike]

    chain.puts["optionType"] = "put"
    chain.calls["optionType"] = "call"

    df = op_helpers.get_greeks(
        current_price=current_price,
        expire=expiry,
        calls=chain.calls,
        puts=chain.puts,
        div_cont=div_cont,
        rf=rf,
        opt_type=opt_type,
        show_extra_greeks=show_all,
    )

    column_formatting = [".1f", ".4f"] + [".6f"] * 4

    if show_all:
        additional_columns = ["Rho", "Phi", "Charm", "Vanna", "Vomma"]
        column_formatting += [".6f"] * len(additional_columns)

    print_rich_table(
        df,
        headers=list(df.columns),
        show_index=False,
        title=f"{symbol} Greeks",
        floatfmt=column_formatting,
    )

    return None
