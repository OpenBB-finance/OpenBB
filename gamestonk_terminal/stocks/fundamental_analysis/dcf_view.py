""" DCF View """
__docformat__ = "numpy"

from typing import List, Union
from datetime import datetime
from pathlib import Path
import random
import os

from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl import Workbook, worksheet
from openpyxl.styles import Font
from sklearn.linear_model import LinearRegression
from bs4 import BeautifulSoup
import regex as re
import yfinance as yf
import pandas as pd
import numpy as np
import requests

from gamestonk_terminal.stocks.fundamental_analysis import dcf_model
from gamestonk_terminal.helper_funcs import get_rf

# pylint: disable=R0902
# pylint: disable=R0912
# pylint: disable=C0302
# pylint: disable=R0915

int_or_str = Union[int, str]


class CreateExcelFA:
    def __init__(self, ticker: str, audit: bool):
        self.audit: bool = audit
        self.wb: Workbook = Workbook()
        self.ws1: worksheet = self.wb.active
        self.ws2: worksheet = self.wb.create_sheet("Free Cash Flows")
        self.ws3: worksheet = self.wb.create_sheet("Explanations")
        self.ws4: worksheet = self.wb.create_sheet("Ratios")
        self.ws1.title = "Financials"
        self.ticker: str = ticker
        self.now: str = datetime.now().strftime("%Y-%m-%d")
        self.letter: int = 0
        self.is_start: int = 4
        self.bs_start: int = 18
        self.cf_start: int = 47
        self.len_data: int = 0
        self.len_pred: int = 10
        self.years: List[str] = []
        self.rounding: int = 0
        self.df_bs: pd.DataFrame = self.get_data("BS", self.bs_start, False)
        self.df_is: pd.DataFrame = self.get_data("IS", self.is_start, True)
        self.df_cf: pd.DataFrame = self.get_data("CF", self.cf_start, False)
        self.info: pd.DataFrame = yf.Ticker(ticker).info
        self.t_bill: float = get_rf()
        self.r_ff: float = dcf_model.get_fama_coe(self.ticker)
        self.sisters: List[str] = dcf_model.others_in_sector(
            self.ticker, self.info["sector"], self.info["industry"]
        )
        self.sister_data: List[List[pd.DataFrame]] = [[pd.DataFrame()]]

    def create_workbook(self):
        self.ws1.column_dimensions["A"].width = 25
        self.ws2.column_dimensions["A"].width = 22
        for column in dcf_model.letters[1:21]:
            self.ws1.column_dimensions[column].width = 14
        for column in dcf_model.letters[1:21]:
            self.ws2.column_dimensions[column].width = 14

        self.ws3.column_dimensions["A"].width = 3
        self.create_header(self.ws1)
        self.create_header(self.ws2)
        self.create_header(self.ws3)
        self.create_header(self.ws4)
        self.add_estimates()
        self.create_dcf()
        if self.audit:
            self.run_audit()

        trypath = os.path.join(
            os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..")),
            "exports",
            "stocks",
            "fundamental_analysis",
            f"{self.ticker} {self.now}.xlsx",
        )

        my_file = Path(trypath)
        if my_file.is_file():
            print("Analysis already ran. Please move file to rerun.")
        else:
            self.wb.save(trypath)
            print(
                f"Analysis ran for {self.ticker}\nPlease look in {trypath} for the file.\n"
            )

    def get_data(self, statement: str, row: int, header: bool) -> pd.DataFrame:
        URL = f"https://stockanalysis.com/stocks/{self.ticker}/financials/"
        if statement == "BS":
            URL += "balance-sheet/"
            title = "Balance Sheet"
            ignores = dcf_model.non_gaap_bs
        if statement == "CF":
            URL += "cash-flow-statement/"
            title = "Cash Flows"
            ignores = dcf_model.non_gaap_cf
        if statement == "IS":
            title = "Income Statement"
            ignores = dcf_model.non_gaap_is

        r = requests.get(URL, headers=dcf_model.headers)

        if "404 - Page Not Found" in r.text:
            raise ValueError("The ticker given is not in the stock analysis website.")
        soup = BeautifulSoup(r.content, "html.parser")

        table = soup.find("table", attrs={"class": re.compile("fintbl")})
        head = table.find("thead")
        columns = head.find_all("th")

        if self.years == []:
            self.years = [
                x.get_text().strip() for x in columns if "-" not in x.get_text().strip()
            ]
            self.len_data = len(self.years) - 1

        if self.rounding == 0:
            phrase = (
                soup.find(
                    "div", attrs={"class": "text-sm text-gray-600 block lg:hidden"}
                )
                .get_text()
                .lower()
            )

            if "thousand" in phrase:
                self.rounding = 1_000
            elif "millions" in phrase:
                self.rounding = 1_000_000
            elif "billions" in phrase:
                self.rounding = 1_000_000_000
            else:
                raise ValueError(
                    "Stock Analysis did not specify a proper rounding amount"
                )

        body = table.find("tbody")
        rows = body.find_all("tr")

        all_data = [
            [
                x.get_text().strip() if x.get_text().strip() != "-" else "0"
                for x in y.find_all("td")
            ]
            for y in rows
        ]

        df = pd.DataFrame(data=all_data)
        df = df.loc[:, ~(df == "Upgrade").any()]
        df = df.set_index(0)
        n = df.shape[1] - self.len_data
        if n > 0:
            df = df.iloc[:, :-n]
        df.columns = self.years[1:]

        for ignore in ignores:
            if ignore in df.index:
                df = df.drop([ignore])
        df = df[df.columns[::-1]]

        self.ws1[f"A{row}"] = title
        self.ws1[f"A{row}"].font = dcf_model.bold_font

        if statement == "IS":
            vals = ["Revenue", dcf_model.gaap_is]
        elif statement == "BS":
            vals = ["Cash & Equivalents", dcf_model.gaap_bs]
        elif statement == "CF":
            vals = ["Net Income", dcf_model.gaap_cf]

        if vals[0] in df.index:
            blank_list = ["0" for _ in df.loc[vals[0]].to_list()]
        else:
            raise ValueError("Dataframe does not have key information.")
        for i, _ in enumerate(vals[1][1:]):
            df = dcf_model.insert_row(vals[1][i + 1], vals[1][i], df, blank_list)

        rowI = row + 1
        names = df.index.values.tolist()

        for name in names:
            self.ws1[f"A{rowI}"] = name
            if name in dcf_model.sum_rows:
                length = self.len_data + (self.len_pred if statement != "CF" else 0)
                for i in range(length):
                    if statement == "CF" and name == "Net Income":
                        pass
                    else:
                        self.ws1[
                            f"{dcf_model.letters[i+1]}{rowI}"
                        ].font = dcf_model.bold_font
                        self.ws1[
                            f"{dcf_model.letters[i+1]}{rowI}"
                        ].border = dcf_model.thin_border_top
            rowI += 1

        column = 1

        for key, value in df.iteritems():
            rowI = row
            if header:
                dcf_model.set_cell(
                    self.ws1,
                    f"{dcf_model.letters[column]}{rowI}",
                    float(key),
                    font=dcf_model.bold_font,
                )
            for item in value:
                rowI += 1
                m = 0 if item is None else float(item.replace(",", ""))
                dcf_model.set_cell(
                    self.ws1,
                    f"{dcf_model.letters[column]}{rowI}",
                    m,
                    num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
                )
            column += 1

        return df

    def add_estimates(self):
        last_year = self.years[1]
        col = self.len_data + 1
        for i in range(self.len_pred):
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col+i]}4",
                int(last_year) + 1 + i,
                font=dcf_model.bold_font,
            )

        for i in range(41):
            col = self.len_pred + self.len_data + 3
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col]}{3+i}",
                fill=dcf_model.green_bg,
                border=dcf_model.thin_border_nr,
            )
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col+1]}{3+i}",
                fill=dcf_model.green_bg,
                border=dcf_model.thin_border_nl,
            )

        dcf_model.set_cell(
            self.ws1,
            f"{dcf_model.letters[col]}3",
            "Linear model",
            alignment=dcf_model.center,
        )
        self.ws1.merge_cells(f"{dcf_model.letters[col]}3:{dcf_model.letters[col+1]}3")
        dcf_model.set_cell(self.ws1, f"{dcf_model.letters[col]}4", "m")
        dcf_model.set_cell(self.ws1, f"{dcf_model.letters[col+1]}4", "b")
        self.get_linear("Date", "Revenue")
        self.get_linear("Revenue", "Cost of Revenue")
        self.get_sum("Gross Profit", "Revenue", [], ["Cost of Revenue"])
        self.get_linear("Revenue", "Selling, General & Admin", True)
        self.get_linear("Revenue", "Research & Development", True)
        self.get_linear("Revenue", "Other Operating Expenses")
        self.get_sum(
            "Operating Income",
            "Gross Profit",
            [],
            [
                "Selling, General & Admin",
                "Research & Development",
                "Other Operating Expenses",
            ],
        )
        self.get_linear("Revenue", "Preferred Dividends")
        self.get_linear("Revenue", "Interest Expense / Income")
        self.get_linear("Revenue", "Other Expense / Income")
        self.get_linear("Operating Income", "Income Tax")
        self.get_sum(
            "Net Income",
            "Operating Income",
            [],
            ["Interest Expense / Income", "Other Expense / Income", "Income Tax"],
        )
        self.custom_exp(
            "Preferred Dividends",
            "Preferred Dividends are not important in a DCF so we do not attempt to predict them.",
        )
        self.get_linear("Revenue", "Cash & Equivalents", True)
        self.get_linear("Revenue", "Short-Term Investments", True)
        self.get_sum(
            "Cash & Cash Equivalents",
            "Cash & Equivalents",
            ["Short-Term Investments"],
            [],
        )
        self.get_linear("Revenue", "Receivables", True)
        self.get_linear("Revenue", "Inventory", True)
        self.get_linear("Revenue", "Other Current Assets")
        self.get_sum(
            "Total Current Assets",
            "Cash & Cash Equivalents",
            ["Receivables", "Inventory", "Other Current Assets"],
            [],
        )
        self.get_linear("Revenue", "Property, Plant & Equipment", True)
        self.get_linear("Revenue", "Long-Term Investments", True)
        self.get_linear("Revenue", "Goodwill and Intangibles", True)
        self.get_linear("Revenue", "Other Long-Term Assets")
        self.get_sum(
            "Total Long-Term Assets",
            "Property, Plant & Equipment",
            [
                "Long-Term Investments",
                "Goodwill and Intangibles",
                "Other Long-Term Assets",
            ],
            [],
        )
        self.get_sum(
            "Total Assets", "Total Current Assets", ["Total Long-Term Assets"], []
        )
        self.get_linear("Revenue", "Accounts Payable")
        self.get_linear("Revenue", "Deferred Revenue")
        self.get_linear("Revenue", "Current Debt")
        self.get_linear("Revenue", "Other Current Liabilities")
        self.get_sum(
            "Total Current Liabilities",
            "Accounts Payable",
            ["Deferred Revenue", "Current Debt", "Other Current Liabilities"],
            [],
        )
        self.get_sum(
            "Long-Term Debt",
            "Total Assets",
            [],
            [
                "Total Current Liabilities",
                "Other Long-Term Liabilities",
                "Shareholders' Equity",
            ],
            text=(
                "This is the plug. For more information on plugs visit https://corporatefina"
                "nceinstitute.com/resources/questions/model-questions/financial-modeling-plug/"
            ),
        )  # This is the plug
        self.get_linear("Revenue", "Other Long-Term Liabilities")
        self.get_sum(
            "Total Long-Term Liabilities",
            "Long-Term Debt",
            ["Other Long-Term Liabilities"],
            [],
        )
        self.get_sum(
            "Total Liabilities",
            "Total Current Liabilities",
            ["Total Long-Term Liabilities"],
            [],
        )
        self.get_linear("Revenue", "Common Stock")
        col = self.len_data + 1
        rer = self.title_to_row("Retained Earnings")
        nir = self.title_to_row("Net Income")
        for i in range(self.len_pred):
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col+i]}{rer}",
                f"={dcf_model.letters[col+i]}{nir}+{dcf_model.letters[col+i-1]}{rer}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

        self.get_linear("Revenue", "Comprehensive Income")
        self.get_sum(
            "Shareholders' Equity",
            "Common Stock",
            ["Retained Earnings", "Comprehensive Income"],
            [],
        )
        self.get_sum(
            "Total Liabilities and Equity",
            "Total Liabilities",
            ["Shareholders' Equity"],
            [],
        )

        dcf_model.set_cell(
            self.ws1,
            "A65",
            (
                "Warning: Stock Analysis does not have all of the cash flow items included. Operating"
                ", Financing, and Investing Cash Flows may not add up to total cash flows."
            ),
            font=dcf_model.red,
        )

    def create_dcf(self):
        self.ws2["A5"] = "Net Income"
        self.ws2["A6"] = "Change in NWC"
        self.ws2["A7"] = "Change in Capex"
        self.ws2["A8"] = "Preferred Dividends"
        self.ws2["A9"] = "Free Cash Flows"
        r = 4

        c1 = dcf_model.letters[self.len_pred + 3]
        c2 = dcf_model.letters[self.len_pred + 4]
        c3 = dcf_model.letters[self.len_pred + 5]
        for i in range(self.len_pred):
            j = 1 + i + self.len_data
            cols = dcf_model.letters
            dcf_model.set_cell(
                self.ws2,
                f"{cols[1+i]}4",
                f"=Financials!{cols[j]}4",
                font=dcf_model.bold_font,
            )
            dcf_model.set_cell(
                self.ws2,
                f"{cols[1+i]}5",
                f"=Financials!{cols[j]}{self.title_to_row('Net Income')}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

            dcf_model.set_cell(
                self.ws2,
                f"{dcf_model.letters[1+i]}6",
                (
                    f"=Financials!{cols[j]}{self.title_to_row('Total Current Assets')}"
                    f"-Financials!{cols[j-1]}{self.title_to_row('Total Current Assets')}"
                    f"-Financials!{cols[j]}{self.title_to_row('Total Current Liabilities')}"
                    f"+Financials!{cols[j-1]}{self.title_to_row('Total Current Liabilities')}"
                ),
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
            dcf_model.set_cell(
                self.ws2,
                f"{dcf_model.letters[1+i]}7",
                (
                    f"=Financials!{cols[j]}{self.title_to_row('Total Long-Term Assets')}"
                    f"-Financials!{cols[j-1]}{self.title_to_row('Total Long-Term Assets')}"
                ),
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
            dcf_model.set_cell(
                self.ws2,
                f"{dcf_model.letters[1+i]}8",
                f"=Financials!{cols[j]}{self.title_to_row('Preferred Dividends')}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
            dcf_model.set_cell(
                self.ws2,
                f"{cols[1+i]}9",
                f"={cols[1+i]}5-{cols[1+i]}6-{cols[1+i]}7-{cols[1+i]}8",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
                font=dcf_model.bold_font,
                border=dcf_model.thin_border_top,
            )
            dcf_model.set_cell(
                self.ws2,
                f"{cols[1+self.len_pred]}9",
                f"=({cols[self.len_pred]}9*(1+{c2}" f"{r+15}))/({c2}{r+11}-{c2}{r+15})",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

        dcf_model.set_cell(
            self.ws2,
            f"{c1}{r-2}",
            "Note: We do not allow r values to go below 0.5%.",
            font=dcf_model.red,
        )
        self.ws2.merge_cells(f"{c1}{r}:{c2}{r}")
        for x in [c1, c2]:
            dcf_model.set_cell(self.ws2, f"{x}{r}", border=dcf_model.thin_border_bottom)
        dcf_model.set_cell(
            self.ws2,
            f"{c1}{r}",
            "Discount Rate",
            alignment=dcf_model.center,
            border=dcf_model.thin_border_bottom,
        )

        # CAPM
        dcf_model.set_cell(self.ws2, f"{c1}{r+1}", "Risk Free Rate")
        dcf_model.set_cell(
            self.ws2,
            f"{c2}{r+1}",
            float(self.t_bill) / 100,
            num_form=FORMAT_PERCENTAGE_00,
        )
        self.custom_exp(r + 1, "Pulled from US Treasurey.", 2, f"{c3}")
        dcf_model.set_cell(self.ws2, f"{c1}{r+2}", "Market Rate")
        dcf_model.set_cell(self.ws2, f"{c2}{r+2}", 0.08, num_form=FORMAT_PERCENTAGE_00)
        self.custom_exp(
            r + 2, "Average return of the S&P 500 is 8% [Investopedia]", 2, f"{c3}"
        )
        dcf_model.set_cell(self.ws2, f"{c1}{r+3}", "Beta")
        if self.info["beta"] is None:
            dcf_model.set_cell(self.ws2, f"{c2}{r+3}", float(1))
            self.custom_exp(
                r + 3, "Warning: Beta not found. Assumed a beta of one.", 2, f"{c3}"
            )
            self.info["beta"] = 1
        else:
            dcf_model.set_cell(self.ws2, f"{c2}{r+3}", float(self.info["beta"]))
            self.custom_exp(r + 3, "Beta from yahoo finance", 2, f"{c3}")
        dcf_model.set_cell(self.ws2, f"{c1}{r+4}", "r (CAPM)")
        dcf_model.set_cell(
            self.ws2,
            f"{c2}{r+4}",
            f"=max((({c2}{r+2}-{c2}{r+1})*{c2}{r+3})+{c2}{r+1},0.005)",
            num_form=FORMAT_PERCENTAGE_00,
            border=dcf_model.thin_border_top,
            font=dcf_model.bold_font,
        )

        # Fama French
        dcf_model.set_cell(self.ws2, f"{c1}{r+7}", "Fama French")
        dcf_model.set_cell(
            self.ws2,
            f"{c2}{r+7}",
            f"=max({self.r_ff},0.005)",
            num_form=FORMAT_PERCENTAGE_00,
        )
        self.custom_exp(
            r + 7,
            (
                "Calculated using the Fama and French Three-Factor model. For more"
                "information visit https://www.investopedia.com/terms/f/famaandfrenchthreefactormodel.asp."
            ),
            2,
            f"{c3}",
        )

        # Decide
        for x in [c1, c2]:
            dcf_model.set_cell(
                self.ws2, f"{x}{r+9}", border=dcf_model.thin_border_bottom
            )
        self.ws2.merge_cells(f"{c1}{r+9}:{c2}{r+9}")
        dcf_model.set_cell(
            self.ws2,
            f"{c1}{r+9}",
            "Choose model",
            border=dcf_model.thin_border_bottom,
            alignment=dcf_model.center,
            num_form=FORMAT_PERCENTAGE_00,
        )
        dcf_model.set_cell(self.ws2, f"{c1}{r+10}", "Model")
        dcf_model.set_cell(self.ws2, f"{c2}{r+10}", "Fama French")
        dcf_model.set_cell(self.ws2, f"{c3}{r+10}", "Type 'Fama French' or 'CAPM'")
        dcf_model.set_cell(self.ws2, f"{c1}{r+11}", "r")
        dcf_model.set_cell(
            self.ws2,
            f"{c2}{r+11}",
            f'=if({c2}{r+10}="Fama French",{c2}{r+7},if({c2}{r+10}="CAPM",{c2}{r+4},"Invalid Selection"))',
            num_form=FORMAT_PERCENTAGE_00,
        )

        dcf_model.set_cell(self.ws2, f"{c1}{r+15}", "Long Term Growth")
        dcf_model.set_cell(
            self.ws2,
            f"{c2}{r+15}",
            f"=min(0.04,{c2}{r+11}*0.9)",
            num_form=FORMAT_PERCENTAGE_00,
        )
        dcf_model.set_cell(self.ws2, "A11", "Value from Operations")
        dcf_model.set_cell(
            self.ws2,
            "B11",
            f"=NPV({c2}{r+11},B9:{dcf_model.letters[self.len_pred+1]}9)",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws2, "A12", "Cash and Cash Equivalents")
        dcf_model.set_cell(
            self.ws2,
            "B12",
            f"=financials!{dcf_model.letters[self.len_data]}{self.title_to_row('Cash & Cash Equivalents')}",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws2, "A13", "Intrinsic Value (sum)")
        dcf_model.set_cell(
            self.ws2,
            "B13",
            "=B11+B12",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws2, "A14", "Debt Obligations")
        dcf_model.set_cell(
            self.ws2,
            "B14",
            f"=financials!{dcf_model.letters[self.len_data]}{self.title_to_row('Total Long-Term Liabilities')}",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws2, "A15", "Firm value without debt")
        dcf_model.set_cell(
            self.ws2,
            "B15",
            (
                f"=max(B13-B14,"
                f"Financials!{dcf_model.letters[self.len_data]}{self.title_to_row('Total Assets')}"
                f"-Financials!{dcf_model.letters[self.len_data]}{self.title_to_row('Total Liabilities')})"
            ),
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(
            self.ws2,
            "C15",
            (
                f"=if((B13-B14)>"
                f"(Financials!{dcf_model.letters[self.len_data]}{self.title_to_row('Total Assets')}"
                f"-Financials!{dcf_model.letters[self.len_data]}{self.title_to_row('Total Liabilities')}),"
                '"","Note: Total assets minus total liabilities exceeds projected firm value without debt.'
                ' Value shown is total assets minus total liabilities.")'
            ),
            font=dcf_model.red,
        )
        dcf_model.set_cell(self.ws2, "A16", "Shares Outstanding")
        dcf_model.set_cell(self.ws2, "B16", int(self.info["sharesOutstanding"]))
        dcf_model.set_cell(self.ws2, "A17", "Shares Price")
        dcf_model.set_cell(
            self.ws2,
            "B17",
            f"=(B15*{self.rounding})/B16",
            num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
        )
        dcf_model.set_cell(self.ws2, "A18", "Actual Price")
        dcf_model.set_cell(self.ws2, "B18", float(self.info["regularMarketPrice"]))

        # Create ratios page
        self.ws4.column_dimensions["A"].width = 27
        dcf_model.set_cell(self.ws4, "B4", "Sector:")
        dcf_model.set_cell(self.ws4, "C4", self.info["sector"])

        self.get_sister_dfs()
        self.sister_data.insert(0, [self.ticker, [self.df_bs, self.df_is, self.df_cf]])
        row = 6
        for i in self.sister_data:
            self.ws4.merge_cells(f"A{row}:J{row}")
            dcf_model.set_cell(self.ws4, f"A{row}", i[0], alignment=dcf_model.center)
            dcf_model.set_cell(
                self.ws4,
                f"A{row+1}",
                "Liquidity Ratios",
                border=dcf_model.thin_border,
                font=dcf_model.bold_font,
            )
            dcf_model.set_cell(self.ws4, f"A{row+2}", "Current Ratio")
            dcf_model.set_cell(self.ws4, f"A{row+3}", "Quick Ratio")
            dcf_model.set_cell(
                self.ws4,
                f"A{row+5}",
                "Activity Ratios",
                border=dcf_model.thin_border,
                font=dcf_model.bold_font,
            )
            dcf_model.set_cell(self.ws4, f"A{row+6}", "AR Turnover")
            dcf_model.set_cell(self.ws4, f"A{row+7}", "Days Sales in AR")
            dcf_model.set_cell(self.ws4, f"A{row+8}", "Inventory Turnover")
            dcf_model.set_cell(self.ws4, f"A{row+9}", "Days in Inventory")
            dcf_model.set_cell(self.ws4, f"A{row+10}", "Average Payable Turnover")
            dcf_model.set_cell(self.ws4, f"A{row+11}", "Days of Payables Outstanding")
            dcf_model.set_cell(self.ws4, f"A{row+12}", "Cash Conversion Cycle")
            dcf_model.set_cell(self.ws4, f"A{row+13}", "Asset Turnover")
            dcf_model.set_cell(
                self.ws4,
                f"A{row+15}",
                "Profitability Ratios",
                border=dcf_model.thin_border,
                font=dcf_model.bold_font,
            )
            dcf_model.set_cell(self.ws4, f"A{row+16}", "Profit Margin")
            dcf_model.set_cell(self.ws4, f"A{row+17}", "Return on Assets")
            dcf_model.set_cell(self.ws4, f"A{row+18}", "Return on Equity")
            dcf_model.set_cell(self.ws4, f"A{row+19}", "Return on Sales")
            dcf_model.set_cell(self.ws4, f"A{row+20}", "Gross Margin")
            dcf_model.set_cell(self.ws4, f"A{row+21}", "Operating Cash Flow Ratio")
            dcf_model.set_cell(
                self.ws4,
                f"A{row+23}",
                "Coverage Ratios",
                border=dcf_model.thin_border,
                font=dcf_model.bold_font,
            )
            dcf_model.set_cell(self.ws4, f"A{row+24}", "Debt-to-Equity")
            dcf_model.set_cell(self.ws4, f"A{row+25}", "Total Debt Ratio")
            dcf_model.set_cell(self.ws4, f"A{row+26}", "Equity Multiplier")
            dcf_model.set_cell(self.ws4, f"A{row+27}", "Times Interest Earned")
            dcf_model.set_cell(
                self.ws4,
                f"A{row+29}",
                "Investor Ratios",
                border=dcf_model.thin_border,
                font=dcf_model.bold_font,
            )
            dcf_model.set_cell(self.ws4, f"A{row+30}", "Earnings Per Share")
            dcf_model.set_cell(self.ws4, f"A{row+31}", "Price Earnings Ratio")
            for j in range(len(self.df_bs.columns) - 1):
                lt = dcf_model.letters[j + 1]

                cace1 = float(
                    i[1][0]
                    .at["Cash & Cash Equivalents", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ar0 = float(
                    i[1][0]
                    .at["Receivables", i[1][0].columns[j]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ar1 = float(
                    i[1][0]
                    .at["Receivables", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                inv0 = float(
                    i[1][0]
                    .at["Inventory", i[1][0].columns[j]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                inv1 = float(
                    i[1][0]
                    .at["Inventory", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ca1 = float(
                    i[1][0]
                    .at["Total Current Assets", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ta0 = float(
                    i[1][0]
                    .at["Total Assets", i[1][0].columns[j]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ta1 = float(
                    i[1][0]
                    .at["Total Assets", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ap0 = float(
                    i[1][0]
                    .at["Accounts Payable", i[1][0].columns[j]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ap1 = float(
                    i[1][0]
                    .at["Accounts Payable", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                cl1 = float(
                    i[1][0]
                    .at["Total Current Liabilities", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                tl1 = float(
                    i[1][0]
                    .at["Total Liabilities", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                te0 = float(
                    i[1][0]
                    .at["Shareholders' Equity", i[1][0].columns[j]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                te1 = float(
                    i[1][0]
                    .at["Shareholders' Equity", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                sls1 = float(
                    i[1][1]
                    .at["Revenue", i[1][1].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                cogs1 = float(
                    i[1][1]
                    .at["Cost of Revenue", i[1][1].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                inte1 = float(
                    i[1][1]
                    .at["Interest Expense / Income", i[1][1].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                tax1 = float(
                    i[1][1]
                    .at["Income Tax", i[1][1].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                ni1 = float(
                    i[1][1]
                    .at["Net Income", i[1][1].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                pdiv1 = float(
                    i[1][1]
                    .at["Preferred Dividends", i[1][0].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )
                opcf1 = float(
                    i[1][2]
                    .at["Operating Cash Flow", i[1][2].columns[j + 1]]
                    .replace(",", "")
                    .replace("-", "-0")
                )

                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+1}",
                    int(i[1][0].columns[j + 1]),
                    font=dcf_model.bold_font,
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+2}", "N/A" if cl1 == 0 else ca1 / cl1
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+3}", "N/A" if cl1 == 0 else (cace1 + ar1) / cl1
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+6}",
                    "N/A" if ar0 + ar1 == 0 else sls1 / ((ar0 + ar1) / 2),
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+7}", "N/A" if sls1 == 0 else ar1 / (sls1 / 365)
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+8}",
                    "N/A" if inv0 + inv1 == 0 else cogs1 / ((inv0 + inv1) / 2),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+9}",
                    "N/A" if cogs1 == 0 else inv1 / (cogs1 / 365),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+10}",
                    "N/A" if ap0 + ap1 == 0 else cogs1 / ((ap0 + ap1) / 2),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+11}",
                    "N/A" if cogs1 == 0 else ap1 / (cogs1 / 365),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+12}",
                    "N/A"
                    if cogs1 == 0
                    else (ar1 / (sls1 / 365))
                    + (inv1 / (cogs1 / 365))
                    - (ap1 / (cogs1 / 365)),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+13}",
                    "N/A" if ta0 + ta1 == 0 else sls1 / ((ta0 + ta1) / 2),
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+16}", "N/A" if sls1 == 0 else ni1 / sls1
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+17}",
                    "N/A" if ar0 + ar1 == 0 else ni1 / ((ar0 + ar1) / 2),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+18}",
                    "N/A" if te0 + te1 == 0 else ni1 / ((te0 + te1) / 2),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+19}",
                    "N/A" if sls1 == 0 else (ni1 + inte1 + tax1) / sls1,
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+20}",
                    "N/A" if sls1 == 0 else (sls1 - cogs1) / sls1,
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+21}", "N/A" if cl1 == 0 else opcf1 / cl1
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+24}", "N/A" if te1 == 0 else tl1 / te1
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+25}", "N/A" if ta1 == 0 else tl1 / ta1
                )
                dcf_model.set_cell(
                    self.ws4, f"{lt}{row+26}", "N/A" if te1 == 0 else ta1 / te1
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+27}",
                    "N/A" if inte1 == 0 else (ni1 + inte1 + tax1) / inte1,
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+30}",
                    (ni1 - pdiv1) / float(self.info["sharesOutstanding"]),
                )
                dcf_model.set_cell(
                    self.ws4,
                    f"{lt}{row+31}",
                    float(self.info["previousClose"])
                    / ((ni1 - pdiv1) / float(self.info["sharesOutstanding"])),
                )
            row += 35

    def create_header(self, ws: Workbook):
        for i in range(10):
            dcf_model.set_cell(
                ws, f"{dcf_model.letters[i]}1", border=dcf_model.thin_border
            )

        ws.merge_cells("A1:J1")
        dcf_model.set_cell(
            ws,
            "A1",
            f"Gamestonk Terminal Analysis: {self.ticker.upper()}",
            font=Font(color="04cca8", size=20),
            border=dcf_model.thin_border,
            alignment=dcf_model.center,
        )
        dcf_model.set_cell(ws, "A2", f"DCF for {self.ticker} generated on {self.now}")

    def run_audit(self):
        start = 67
        for i, value in enumerate(dcf_model.sum_rows):
            dcf_model.set_cell(self.ws1, f"A{start + i}", value)

        self.ws1.merge_cells(f"A{start-2}:K{start-2}")
        dcf_model.set_cell(
            self.ws1,
            f"A{start - 2}",
            "Financial Statement Audit",
            font=Font(color="FF0000"),
            alignment=dcf_model.center,
        )
        dcf_model.set_cell(
            self.ws1,
            f"A{start - 1}",
            "Audit ensures data integrity. Numbers should be 0 (with slight rounding difference).",
        )

        self.get_sum(start, "Revenue", [], ["Cost of Revenue", "Gross Profit"], True)
        self.get_sum(
            start + 1,
            "Gross Profit",
            [],
            [
                "Selling, General & Admin",
                "Research & Development",
                "Other Operating Expenses",
                "Operating Income",
            ],
            True,
        )
        self.get_sum(
            start + 2,
            "Operating Income",
            [],
            [
                "Interest Expense / Income",
                "Other Expense / Income",
                "Income Tax",
                "Net Income",
            ],
            True,
        )
        self.get_sum(
            start + 3,
            "Cash & Equivalents",
            ["Short-Term Investments"],
            ["Cash & Cash Equivalents"],
            True,
        )
        self.get_sum(
            start + 4,
            "Cash & Cash Equivalents",
            ["Receivables", "Inventory", "Other Current Assets"],
            ["Total Current Assets"],
            True,
        )
        self.get_sum(
            start + 5,
            "Property, Plant & Equipment",
            [
                "Long-Term Investments",
                "Goodwill and Intangibles",
                "Other Long-Term Assets",
            ],
            ["Total Long-Term Assets"],
            True,
        )
        self.get_sum(
            start + 6,
            "Total Current Assets",
            ["Total Long-Term Assets"],
            ["Total Assets"],
            True,
        )
        self.get_sum(
            start + 7,
            "Accounts Payable",
            ["Deferred Revenue", "Current Debt", "Other Current Liabilities"],
            ["Total Current Liabilities"],
            True,
        )
        self.get_sum(
            start + 8,
            "Long-Term Debt",
            ["Other Long-Term Liabilities"],
            ["Total Long-Term Liabilities"],
            True,
        )
        self.get_sum(
            start + 9,
            "Total Current Liabilities",
            ["Total Long-Term Liabilities"],
            ["Total Liabilities"],
            True,
        )
        self.get_sum(
            start + 10,
            "Common Stock",
            ["Retained Earnings", "Comprehensive Income"],
            ["Shareholders' Equity"],
            True,
        )
        self.get_sum(
            start + 11,
            "Total Liabilities",
            ["Shareholders' Equity"],
            ["Total Liabilities and Equity"],
            True,
        )
        self.get_sum(
            start + 12,
            "Net Income",
            [
                "Depreciation & Amortization",
                "Share-Based Compensation",
                "Other Operating Activities",
            ],
            ["Operating Cash Flow"],
            True,
        )
        self.get_sum(
            start + 13,
            "Capital Expenditures",
            ["Acquisitions", "Change in Investments", "Other Investing Activities"],
            ["Investing Cash Flow"],
            True,
        )
        self.get_sum(
            start + 14,
            "Dividends Paid",
            [
                "Share Issuance / Repurchase",
                "Debt Issued / Paid",
                "Other Financing Activities",
            ],
            ["Financing Cash Flow"],
            True,
        )
        self.get_sum(
            start + 15,
            "Operating Cash Flow",
            ["Investing Cash Flow", "Financing Cash Flow"],
            ["Net Cash Flow"],
            True,
        )

    def get_linear(self, x_ind: str, y_ind: str, no_neg: bool = False):
        x_type = "IS" if x_ind in self.df_is.index else "BS"
        y_type = "IS" if y_ind in self.df_is.index else "BS"
        x_df = self.df_is if x_type == "IS" else self.df_bs
        y_df = self.df_is if y_type == "IS" else self.df_bs
        pre_x = (
            x_df.columns.to_numpy() if x_ind == "Date" else x_df.loc[x_ind].to_numpy()
        )

        vfunc = np.vectorize(dcf_model.string_float)
        pre_x = vfunc(pre_x)

        if x_ind == "Date":
            pre_x = pre_x - np.min(pre_x)
        x = pre_x.reshape((-1, 1))
        pre_y = y_df.loc[y_ind].to_numpy()
        y = vfunc(pre_y)
        model = LinearRegression().fit(x, y)
        r_sq = model.score(x, y)
        r = abs(r_sq ** 0.5)

        if r > 0.9:
            strength = "very strong"
        elif r > 0.7:
            strength = "strong"
        elif r > 0.5:
            strength = "moderate"
        elif r > 0.3:
            strength = "weak"
        else:
            strength = "very weak"

        row1 = (
            y_df.index.get_loc(y_ind)
            + 1
            + (self.is_start if y_type == "IS" else self.bs_start)
        )

        col = self.len_pred + self.len_data + 3
        dcf_model.set_cell(
            self.ws1, f"{dcf_model.letters[col]}{row1}", float(model.coef_)
        )
        dcf_model.set_cell(
            self.ws1, f"{dcf_model.letters[col+1]}{row1}", float(model.intercept_)
        )
        dcf_model.set_cell(
            self.ws1,
            f"{dcf_model.letters[col+2]}{row1}",
            dcf_model.letters[self.letter],
            font=dcf_model.red,
        )
        dcf_model.set_cell(
            self.ws3,
            f"A{self.letter+4}",
            dcf_model.letters[self.letter],
            font=dcf_model.red,
        )
        dcf_model.set_cell(
            self.ws3,
            f"B{self.letter+4}",
            (
                f"The correlation between {x_ind.lower()} and {y_ind.lower()}"
                f" is {strength} with a correlation coefficient of {r:.4f}."
            ),
        )

        col = self.len_data + 1
        for i in range(self.len_pred):
            if x_ind == "Date":
                base = (
                    f"(({dcf_model.letters[col+i]}4-B4)*{dcf_model.letters[col+self.len_pred+2]}"
                    f"{row1})+{dcf_model.letters[col+self.len_pred+3]}{row1}"
                )
            else:
                row_n = (
                    x_df.index.get_loc(x_ind) + 1 + self.is_start
                    if x_type == "IS"
                    else self.bs_start
                )
                base = (
                    f"({dcf_model.letters[col+i]}{row_n}*{dcf_model.letters[col+self.len_pred+2]}{row1})"
                    f"+{dcf_model.letters[col+self.len_pred+3]}{row1}"
                )
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col+i]}{row1}",
                f"=max({base},0)" if no_neg else f"={base}",
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )

        self.letter += 1

    def get_sum(
        self,
        row: int_or_str,
        first: str,
        adds: List[str],
        subtracts: List[str],
        audit: bool = False,
        text: str = None,
    ):
        col = 1 if audit else self.len_data + 1
        for i in range(self.len_data if audit else self.len_pred):
            sum_formula = f"={dcf_model.letters[col+i]}{self.title_to_row(first)}"
            for item in adds:
                sum_formula += f"+{dcf_model.letters[col+i]}{self.title_to_row(item)}"
            for item in subtracts:
                sum_formula += f"-{dcf_model.letters[col+i]}{self.title_to_row(item)}"
            rowI = row if isinstance(row, int) else self.title_to_row(row)
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col+i]}{rowI}",
                sum_formula,
                num_form="[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00",
            )
        if text:
            self.custom_exp(row, text)

    def title_to_row(self, title: str) -> int:
        df = (
            self.df_is
            if title in self.df_is.index
            else self.df_bs
            if title in self.df_bs.index
            else self.df_cf
        )
        ind = (
            df.index.get_loc(title)
            + 1
            + (
                self.is_start
                if title in self.df_is.index
                else self.bs_start
                if title in self.df_bs.index
                else self.cf_start
            )
        )
        return ind

    def custom_exp(self, row: int_or_str, text: str, ws: int = 1, column: str = None):
        if ws == 1:
            rowT = row if isinstance(row, int) else self.title_to_row(row)
            col = self.len_pred + self.len_data + 3
            dcf_model.set_cell(
                self.ws1,
                f"{dcf_model.letters[col+2]}{rowT}",
                dcf_model.letters[self.letter],
                font=dcf_model.red,
            )
        if ws == 2:
            dcf_model.set_cell(
                self.ws2,
                f"{column}{row}",
                dcf_model.letters[self.letter],
                font=dcf_model.red,
            )

        dcf_model.set_cell(
            self.ws3,
            f"A{self.letter+4}",
            dcf_model.letters[self.letter],
            font=dcf_model.red,
        )
        dcf_model.set_cell(self.ws3, f"B{self.letter+4}", text)
        self.letter += 1

    def get_sister_dfs(self):
        # TODO: Once mcap is added to this, we can add as an additional filters for more comparative results
        sisters = self.sisters
        random.shuffle(sisters)
        i = 0
        new_list = []
        while i < 3 and sisters:
            sister_ret = [
                self.get_sister_data(x, sisters[0]) for x in ["BS", "IS", "CF"]
            ]
            print(sister_ret)
            blank = [x.empty for x in sister_ret]
            if True in blank:
                sisters.pop(0)
            else:
                vals = [
                    sisters[0],
                    sister_ret,
                ]
                new_list.append(vals)
                i += 1
                sisters.pop(0)

        self.sister_data = new_list

    def get_sister_data(self, statement: str, ticker: str) -> pd.DataFrame:
        URL = f"https://stockanalysis.com/stocks/{ticker}/financials/"
        if statement == "BS":
            URL += "balance-sheet/"
            ignores = dcf_model.non_gaap_bs
        if statement == "CF":
            URL += "cash-flow-statement/"
            ignores = dcf_model.non_gaap_cf
        if statement == "IS":
            ignores = dcf_model.non_gaap_is

        r = requests.get(URL, headers=dcf_model.headers)

        if "404 - Page Not Found" in r.text:
            # TODO: add better handling
            return pd.DataFrame()
        soup = BeautifulSoup(r.content, "html.parser")

        table = soup.find("table", attrs={"class": re.compile("fintbl")})
        head = table.find("thead")
        if head is None:
            return pd.DataFrame()
        columns = head.find_all("th")

        if self.years == []:
            self.years = [
                x.get_text().strip() for x in columns if "-" not in x.get_text().strip()
            ]
            self.len_data = len(self.years) - 1

        if self.rounding == 0:
            phrase = (
                soup.find(
                    "div", attrs={"class": "text-sm text-gray-600 block lg:hidden"}
                )
                .get_text()
                .lower()
            )
            if "thousand" in phrase:
                self.rounding = 1_000
            elif "millions" in phrase:
                self.rounding = 1_000_000
            elif "billions" in phrase:
                self.rounding = 1_000_000_000
            else:
                return pd.DataFrame

        body = table.find("tbody")
        rows = body.find_all("tr")

        all_data = [[x.get_text().strip() for x in y.find_all("td")] for y in rows]

        df = pd.DataFrame(data=all_data)
        df = df.loc[:, ~(df == "Upgrade").any()]
        df = df.set_index(0)
        n = df.shape[1] - self.len_data
        if n > 0:
            df = df.iloc[:, :-n]

        df.columns = self.years[1:]

        for ignore in ignores:
            if ignore in df.index:
                df = df.drop([ignore])
        df = df[df.columns[::-1]]

        if statement == "IS":
            vals = ["Revenue", dcf_model.gaap_is]
        elif statement == "BS":
            vals = ["Cash & Equivalents", dcf_model.gaap_bs]
        elif statement == "CF":
            vals = ["Net Income", dcf_model.gaap_cf]

        if vals[0] in df.index:
            blank_list = ["0" for _ in df.loc[vals[0]].to_list()]
        else:
            return pd.DataFrame()
        for i, _ in enumerate(vals[1][1:]):
            df = dcf_model.insert_row(vals[1][i + 1], vals[1][i], df, blank_list)

        return df
