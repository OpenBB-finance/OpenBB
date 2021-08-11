"""
Known issues:
-The statement of cash flows provided does not tie out.
This is an issue with data from the website and does not affect my calculations.
"""

from datetime import datetime
import math
from typing import List, Literal, Union
import argparse

from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl import Workbook, worksheet
from sklearn.linear_model import LinearRegression
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
import numpy as np
import requests

import gamestonk_terminal.stocks.fundamental_analysis.excel.variables as var
from gamestonk_terminal.helper_funcs import parse_known_args_and_warn

stmt = Literal["IS", "BS", "CF"]
int_or_str = Union[int, str]


def string_float(string: str):
    return float(string.replace(",", ""))


def insert_row(name: str, index: str, df: pd.DataFrame, row_value: List[str]):
    pd.options.mode.chained_assignment = None
    if name not in df.index:
        row_number = df.index.get_loc(index) + 1
        df1 = df[0:row_number]
        df2 = df[row_number:]
        df1.loc[name] = row_value
        df_result = pd.concat([df1, df2])
        return df_result
    return df


def excel(other_args: List[str], ticker: str):
    """Discounted cash flow

    Parameters
    ----------
    other_args : List[str]
        argparse other args
    ticker : str
        Fundamental analysis ticker symbol
    """
    parser = argparse.ArgumentParser(
        add_help=False,
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        prog="excel",
        description="""
            Generates a completed discounted cash flow statement. The statement uses machine
             learning to predict the future financial statement, and then predicts the future
             value of the stock based on the predicted financials.""",
    )
    parser.add_argument(
        "-a",
        "--audit",
        action="store_true",
        dest="audit",
        default=False,
        help="Confirms that the numbers provided are accurate.",
    )

    # try:
    ns_parser = parse_known_args_and_warn(parser, other_args)
    if not ns_parser:
        return

    excel_view = CreateExcelFA(ticker, ns_parser.audit)
    excel_view.create_workbook()

    # except Exception as e:
    # print(e, "\n")


class CreateExcelFA:

    # pylint: disable=R0902
    # pylint: disable=R0912
    # Pylint is mad because I have too many class attributes and methods.
    # If a reviewer thinks I need to refactor please let me know.

    def __init__(self, ticker: str, audit: bool):
        self.audit: bool = audit
        self.wb: Workbook = Workbook()
        self.ws1: worksheet = self.wb.active
        self.ws2: worksheet = self.wb.create_sheet("Free Cash Flows")
        self.ws3: worksheet = self.wb.create_sheet("Explanations")
        self.ws1.title = "Financials"
        self.ticker: str = ticker
        self.now: str = datetime.now().strftime("%m/%d/%Y, %H:%M:%S").replace("/", "-")
        self.letter: int = 0
        self.is_start: int = 4
        self.bs_start: int = 18
        self.cf_start: int = 47
        self.len_data: int = 0
        self.len_pred: int = 10
        self.years: List[str] = []
        self.rounding: int = 0
        self.df_is: pd.DataFrame = self.get_data("IS", self.is_start, True)
        self.df_bs: pd.DataFrame = self.get_data("BS", self.bs_start, False)
        self.df_cf: pd.DataFrame = self.get_data("CF", self.cf_start, False)
        self.info = yf.Ticker(ticker).info

    def create_workbook(self):
        self.ws1.column_dimensions["A"].width = 25
        self.ws2.column_dimensions["A"].width = 22
        for column in var.letters[1:21]:
            self.ws1.column_dimensions[column].width = 14
        for column in var.letters[1:21]:
            self.ws2.column_dimensions[column].width = 14
        for i in range(50):
            if i != 1:
                self.ws3[f"A{str(1+i)}"].font = Font(color="FF0000")
        self.ws3.column_dimensions["A"].width = 3
        self.create_header(self.ws1)
        self.create_header(self.ws2)
        self.create_header(self.ws3)
        self.add_estimates()
        self.create_dcf()
        if self.audit:
            self.run_audit()
        self.wb.save("../" + f"{self.ticker}-{self.now}" + ".xlsx")

    def get_data(self, statement: stmt, row: int, header: bool):
        URL = f"https://stockanalysis.com/stocks/{self.ticker}/financials/"
        if statement == "BS":
            URL += "balance-sheet/"
            title = "Balance Sheet"
            ignores = var.non_gaap_bs
        if statement == "CF":
            URL += "cash-flow-statement/"
            title = "Cash Flows"
            ignores = var.non_gaap_cf
        if statement == "IS":
            title = "Income Statement"
            ignores = var.non_gaap_is

        r = requests.get(URL, headers=var.headers)
        if "404 - Page Not Found" in r.text:
            raise ValueError("The ticker given is not in the stock analysis website.")
        soup = BeautifulSoup(r.content, "lxml")

        table = soup.find(
            "table", attrs={"class": "FinancialTable_table_financial__1RhYq"}
        )
        head = table.find("thead")
        columns = head.find_all("th")

        if self.years == []:
            self.years = [x.get_text().strip() for x in columns]

        if self.rounding == 0:
            phrase = soup.find(
                "div", attrs={"class": "text-sm pb-1 text-gray-600"}
            ).get_text()
            if "thousand" in phrase:
                self.rounding = 1000
            elif "millions" in phrase:
                self.rounding = 1000000
            elif "billions" in phrase:
                self.rounding = 1000000000
            else:
                raise ValueError(
                    "Stock Analysis did not specify a proper rounding amount"
                )

        body = table.find("tbody")
        rows = body.find_all("tr")

        all_data = [[x.get_text().strip() for x in y.find_all("td")] for y in rows]

        df = pd.DataFrame(data=all_data)
        df = df.set_index(0)
        df.columns = self.years[1:]

        for ignore in ignores:
            if ignore in df.index:
                df = df.drop([ignore])
        df = df[df.columns[::-1]]

        self.ws1[f"A{row}"] = title
        self.ws1[f"A{row}"].font = var.bold_font

        if statement == "IS":
            if "Revenue" in df.index:
                blank_list = ["0" for x in df.loc["Revenue"].to_list()]
            else:
                raise ValueError("Dataframe does not have key information.")
            for i, value in enumerate(var.gaap_is[1:]):
                df = insert_row(var.gaap_is[i + 1], var.gaap_is[i], df, blank_list)

        if statement == "BS":
            if "Cash & Equivalents" in df.index:
                blank_list = ["0" for x in df.loc["Cash & Equivalents"].to_list()]
            else:
                raise ValueError("Dataframe does not have key information.")
            for i, value in enumerate(var.gaap_bs[1:]):
                df = insert_row(var.gaap_bs[i + 1], var.gaap_bs[i], df, blank_list)

        if statement == "CF":
            if "Net Income" in df.index:
                blank_list = ["0" for x in df.loc["Net Income"].to_list()]
            else:
                raise ValueError("Dataframe does not have key information.")
            for i, value in enumerate(var.gaap_cf[1:]):
                df = insert_row(var.gaap_cf[i + 1], var.gaap_cf[i], df, blank_list)

        if self.len_data == 0:
            self.len_data = len(df.columns)

        rowI = row + 1
        names = df.index.values.tolist()

        for name in names:
            self.ws1[f"A{rowI}"] = name
            if name in var.sum_rows:
                length = self.len_data + (self.len_pred if statement != "CF" else 0)
                for i in range(length):
                    if statement == "CF" and name == "Net Income":
                        pass
                    else:
                        self.ws1[f"{var.letters[i+1]}{rowI}"].font = var.bold_font
                        self.ws1[
                            f"{var.letters[i+1]}{rowI}"
                        ].border = var.thin_border_top
            rowI += 1

        column = 1
        for key, value in df.iteritems():
            rowI = row
            if header:
                self.ws1[f"{var.letters[column]}{rowI}"] = float(key)
                self.ws1[f"{var.letters[column]}{rowI}"].font = var.bold_font
            for item in value:
                rowI += 1
                self.ws1[f"{var.letters[column]}{rowI}"] = float(item.replace(",", ""))
                self.ws1[f"{var.letters[column]}{rowI}"].number_format = var.fmt_acct
            column += 1

        return df

    def add_estimates(self):
        last_year = self.years[1]
        col = self.len_data + 1
        for i in range(self.len_pred):
            self.ws1[f"{var.letters[col+i]}4"] = int(last_year) + 1 + i
            self.ws1[f"{var.letters[col+i]}4"].font = var.bold_font

        for i in range(41):
            col = self.len_pred + self.len_data + 3
            self.ws1[f"{var.letters[col]}{3+i}"].fill = PatternFill(
                fgColor="7fe5cd", fill_type="solid"
            )
            self.ws1[f"{var.letters[col+1]}{3+i}"].fill = PatternFill(
                fgColor="7fe5cd", fill_type="solid"
            )
            self.ws1[f"{var.letters[col]}{3+i}"].border = var.thin_border_nr
            self.ws1[f"{var.letters[col+1]}{3+i}"].border = var.thin_border_nl

        self.ws1[f"{var.letters[col]}3"] = "Linear model"
        self.ws1[f"{var.letters[col]}3"].alignment = Alignment(horizontal="center")
        self.ws1.merge_cells(f"{var.letters[col]}3:{var.letters[col+1]}3")
        self.ws1[f"{var.letters[col]}4"] = "m"
        self.ws1[f"{var.letters[col+1]}4"] = "b"
        self.get_linear("Date", "Revenue")
        self.get_linear("Revenue", "Cost of Revenue")
        self.get_sum("Gross Profit", "Revenue", [], ["Cost of Revenue"])
        self.get_linear("Revenue", "Selling, General & Admin", True)
        self.get_linear("Revenue", "Research & Development", True)
        self.get_linear("Revenue", "Other Operating Expenses")
        self.get_sum(
            "Operating Income",
            "Gross Profit",
            [],
            [
                "Selling, General & Admin",
                "Research & Development",
                "Other Operating Expenses",
            ],
        )
        self.get_linear("Revenue", "Preferred Dividends")
        self.get_linear("Revenue", "Interest Expense / Income")
        self.get_linear("Revenue", "Other Expense / Income")
        self.get_linear("Operating Income", "Income Tax")
        self.get_sum(
            "Net Income",
            "Operating Income",
            [],
            ["Interest Expense / Income", "Other Expense / Income", "Income Tax"],
        )
        self.custom_exp(
            "Preferred Dividends",
            "Preferred Dividends are not important in a DCF so we do not attempt to predict them.",
        )
        self.get_linear("Revenue", "Cash & Equivalents", True)
        self.get_linear("Revenue", "Short-Term Investments", True)
        self.get_sum(
            "Cash & Cash Equivalents",
            "Cash & Equivalents",
            ["Short-Term Investments"],
            [],
        )
        self.get_linear("Revenue", "Receivables", True)
        self.get_linear("Revenue", "Inventory", True)
        self.get_linear("Revenue", "Other Current Assets")
        self.get_sum(
            "Total Current Assets",
            "Cash & Cash Equivalents",
            ["Receivables", "Inventory", "Other Current Assets"],
            [],
        )
        self.get_linear("Revenue", "Property, Plant & Equipment", True)
        self.get_linear("Revenue", "Long-Term Investments", True)
        self.get_linear("Revenue", "Goodwill and Intangibles", True)
        self.get_linear("Revenue", "Other Long-Term Assets")
        self.get_sum(
            "Total Long-Term Assets",
            "Property, Plant & Equipment",
            [
                "Long-Term Investments",
                "Goodwill and Intangibles",
                "Other Long-Term Assets",
            ],
            [],
        )
        self.get_sum(
            "Total Assets", "Total Current Assets", ["Total Long-Term Assets"], []
        )
        self.get_linear("Revenue", "Accounts Payable")
        self.get_linear("Revenue", "Deferred Revenue")
        self.get_linear("Revenue", "Current Debt")
        self.get_linear("Revenue", "Other Current Liabilities")
        self.get_sum(
            "Total Current Liabilities",
            "Accounts Payable",
            ["Deferred Revenue", "Current Debt", "Other Current Liabilities"],
            [],
        )
        self.get_sum(
            "Long-Term Debt",
            "Total Assets",
            [],
            [
                "Total Current Liabilities",
                "Other Long-Term Liabilities",
                "Shareholders' Equity",
            ],
            text=(
                "This is the plug. For more information on plugs visit https://corporatefina"
                "nceinstitute.com/resources/questions/model-questions/financial-modeling-plug/"
            ),
        )  # This is the plug
        self.get_linear("Revenue", "Other Long-Term Liabilities")
        self.get_sum(
            "Total Long-Term Liabilities",
            "Long-Term Debt",
            ["Other Long-Term Liabilities"],
            [],
        )
        self.get_sum(
            "Total Liabilities",
            "Total Current Liabilities",
            ["Total Long-Term Liabilities"],
            [],
        )
        self.get_linear("Revenue", "Common Stock")
        col = self.len_data + 1
        rer = self.title_to_row("Retained Earnings")
        nir = self.title_to_row("Net Income")
        for i in range(self.len_pred):
            self.ws1[
                f"{var.letters[col+i]}{rer}"
            ] = f"={var.letters[col+i]}{nir}+{var.letters[col+i-1]}{rer}"
            self.ws1[f"{var.letters[col+i]}{rer}"].number_format = var.fmt_acct
        self.get_linear("Revenue", "Comprehensive Income")
        self.get_sum(
            "Shareholders' Equity",
            "Common Stock",
            ["Retained Earnings", "Comprehensive Income"],
            [],
        )
        self.get_sum(
            "Total Liabilities and Equity",
            "Total Liabilities",
            ["Shareholders' Equity"],
            [],
        )

    def create_dcf(self):
        self.ws2["A5"] = "Net Income"
        self.ws2["A6"] = "Change in NWC"
        self.ws2["A7"] = "Change in Capex"
        self.ws2["A8"] = "Preferred Dividends"
        self.ws2["A9"] = "Free Cash Flows"
        r = 4
        c1 = var.letters[self.len_data + 3]
        c2 = var.letters[self.len_data + 4]
        c3 = var.letters[self.len_data + 5]
        for i in range(self.len_pred):
            self.ws2[
                f"{var.letters[1+i]}4"
            ] = f"=Financials!{var.letters[1+i+self.len_data]}4"
            self.ws2[f"{var.letters[1+i]}4"].font = var.bold_font
            self.ws2[
                f"{var.letters[1+i]}5"
            ] = f"=Financials!{var.letters[1+i+self.len_data]}{self.title_to_row('Net Income')}"
            self.ws2[f"{var.letters[1+i]}5"].number_format = var.fmt_acct
            self.ws2[f"{var.letters[1+i]}6"] = (
                f"=Financials!{var.letters[1+i+self.len_data]}{self.title_to_row('Total Current Assets')}"
                f"-Financials!{var.letters[1+i+self.len_data-1]}{self.title_to_row('Total Current Assets')}"
                f"-Financials!{var.letters[1+i+self.len_data]}{self.title_to_row('Total Current Liabilities')}"
                f"+Financials!{var.letters[1+i+self.len_data-1]}{self.title_to_row('Total Current Liabilities')}"
            )
            self.ws2[f"{var.letters[1+i]}6"].number_format = var.fmt_acct
            self.ws2[f"{var.letters[1+i]}7"] = (
                f"=Financials!{var.letters[1+i+self.len_data]}{self.title_to_row('Total Long-Term Assets')}"
                f"-Financials!{var.letters[1+i+self.len_data-1]}{self.title_to_row('Total Long-Term Assets')}"
            )
            self.ws2[f"{var.letters[1+i]}7"].number_format = var.fmt_acct
            self.ws2[
                f"{var.letters[1+i]}8"
            ] = f"=Financials!{var.letters[1+i+self.len_data]}{self.title_to_row('Preferred Dividends')}"
            self.ws2[f"{var.letters[1+i]}8"].number_format = var.fmt_acct
            self.ws2[
                f"{var.letters[1+i]}9"
            ] = f"={var.letters[1+i]}5-{var.letters[1+i]}6-{var.letters[1+i]}7-{var.letters[1+i]}8"
            self.ws2[f"{var.letters[1+i]}9"].number_format = var.fmt_acct
            self.ws2[f"{var.letters[1+i]}9"].font = var.bold_font
            self.ws2[f"{var.letters[1+i]}9"].border = var.thin_border_top

        self.ws2[f"{var.letters[1+self.len_pred]}9"] = (
            f"=({var.letters[self.len_pred]}9*(1+{c2}" f"{r+6}))/({c2}{r+4}-{c2}{r+6})"
        )

        self.ws2.merge_cells(f"{c1}{r}:{c2}{r}")
        self.ws2[f"{c1}{r}"] = "Discount Rate"
        self.ws2[f"{c1}{r}"].alignment = Alignment(horizontal="center")
        self.ws2[f"{c1}{r+1}"] = "Risk Free Rate"
        self.ws2[f"{c2}{r+1}"] = 0.02
        self.ws2[f"{c2}{r+1}"].number_format = FORMAT_PERCENTAGE_00
        self.ws2[f"{c3}{r+1}"] = "Eventually get from 10 year t-bond scraper"
        self.ws2[f"{c1}{r+2}"] = "Market Rate"
        self.ws2[f"{c2}{r+2}"] = 0.08
        self.ws2[f"{c2}{r+2}"].number_format = FORMAT_PERCENTAGE_00
        self.custom_exp(
            r + 2, "Average return of the S&P 500 is 8% [Investopedia]", 2, f"{c3}"
        )
        self.ws2[f"{c1}{r+3}"] = "Beta"
        self.ws2[f"{c2}{r+3}"] = float(self.info["beta"])
        self.custom_exp(r + 3, "Beta from yahoo finance", 2, f"{c3}")
        self.ws2[f"{c1}{r+4}"] = "r"
        self.ws2[f"{c2}{r+4}"] = f"=(({c2}{r+2}-{c2}{r+1})*{c2}{r+3})+{c2}{r+1}"
        self.ws2[f"{c2}{r+4}"].number_format = FORMAT_PERCENTAGE_00
        self.ws2[f"{c2}{r+4}"].border = var.thin_border_top
        self.ws2[f"{c2}{r+4}"].font = var.bold_font
        self.ws2[f"{c1}{r+6}"] = "Long Term Growth"
        self.ws2[f"{c2}{r+6}"] = 0.04
        self.ws2[f"{c2}{r+6}"].number_format = FORMAT_PERCENTAGE_00

        self.ws2["A11"] = "Value from Operations"
        self.ws2["B11"] = f"=NPV({c2}{r+4},B9:{var.letters[self.len_pred+1]}9)"
        self.ws2["B11"].number_format = var.fmt_acct
        self.ws2["A12"] = "Cash and Cash Equivalents"
        self.ws2[
            "B12"
        ] = f"=financials!{var.letters[self.len_data]}{self.title_to_row('Cash & Cash Equivalents')}"
        self.ws2["B12"].number_format = var.fmt_acct
        self.ws2["A13"] = "Intrinsic Value (sum)"
        self.ws2["B13"] = "=B11+B12"
        self.ws2["B13"].number_format = var.fmt_acct
        self.ws2["A14"] = "Debt Obligations"
        self.ws2[
            "B14"
        ] = f"=financials!{var.letters[self.len_data]}{self.title_to_row('Total Long-Term Liabilities')}"
        self.ws2["B14"].number_format = var.fmt_acct
        self.ws2["A15"] = "Firm value without debt"
        self.ws2["B15"] = "=B13-B14"
        self.ws2["B15"].number_format = var.fmt_acct
        self.ws2["A16"] = "Shares Outstanding"
        self.ws2["B16"] = int(self.info["sharesOutstanding"])
        self.ws2["A17"] = "Shares Price"
        self.ws2["B17"] = f"=(B15*{self.rounding})/B16"
        self.ws2["B17"].number_format = var.fmt_acct

    def create_header(self, ws: Workbook):
        for cell in ws["A1:J1"][0]:
            cell.border = var.thin_border
        ws.merge_cells("A1:J1")
        ws["A1"] = f"Gamestonk Terminal Analysis: {self.ticker.upper()}"
        ws["A1"].font = Font(color="04cca8", size=20)
        ws["A1"].border = var.thin_border
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = f"DCF for {self.ticker} generated on {self.now}"

    def run_audit(self):
        start = 67
        for i, value in enumerate(var.sum_rows):
            self.ws1[f"A{start + i}"] = value

        self.ws1.merge_cells(f"A{start-2}:K{start-2}")
        self.ws1[f"A{start - 2}"].font = Font(color="FF0000")
        self.ws1[f"A{start - 2}"].alignment = Alignment(horizontal="center")
        self.ws1[f"A{start - 2}"] = "Financial Statement Audit"
        self.ws1[
            f"A{start - 1}"
        ] = "This report ensures data integrity. Numbers should be 0 (there may be a slight difference due to rounding)."

        self.get_sum(start, "Revenue", [], ["Cost of Revenue", "Gross Profit"], True)
        self.get_sum(
            start + 1,
            "Gross Profit",
            [],
            [
                "Selling, General & Admin",
                "Research & Development",
                "Other Operating Expenses",
                "Operating Income",
            ],
            True,
        )
        self.get_sum(
            start + 2,
            "Operating Income",
            [],
            [
                "Interest Expense / Income",
                "Other Expense / Income",
                "Income Tax",
                "Net Income",
            ],
            True,
        )
        self.get_sum(
            start + 3,
            "Cash & Equivalents",
            ["Short-Term Investments"],
            ["Cash & Cash Equivalents"],
            True,
        )
        self.get_sum(
            start + 4,
            "Cash & Cash Equivalents",
            ["Receivables", "Inventory", "Other Current Assets"],
            ["Total Current Assets"],
            True,
        )
        self.get_sum(
            start + 5,
            "Property, Plant & Equipment",
            [
                "Long-Term Investments",
                "Goodwill and Intangibles",
                "Other Long-Term Assets",
            ],
            ["Total Long-Term Assets"],
            True,
        )
        self.get_sum(
            start + 6,
            "Total Current Assets",
            ["Total Long-Term Assets"],
            ["Total Assets"],
            True,
        )
        self.get_sum(
            start + 7,
            "Accounts Payable",
            ["Deferred Revenue", "Current Debt", "Other Current Liabilities"],
            ["Total Current Liabilities"],
            True,
        )
        self.get_sum(
            start + 8,
            "Long-Term Debt",
            ["Other Long-Term Liabilities"],
            ["Total Long-Term Liabilities"],
            True,
        )
        self.get_sum(
            start + 9,
            "Total Current Liabilities",
            ["Total Long-Term Liabilities"],
            ["Total Liabilities"],
            True,
        )
        self.get_sum(
            start + 10,
            "Common Stock",
            ["Retained Earnings", "Comprehensive Income"],
            ["Shareholders' Equity"],
            True,
        )
        self.get_sum(
            start + 11,
            "Total Liabilities",
            ["Shareholders' Equity"],
            ["Total Liabilities and Equity"],
            True,
        )
        self.get_sum(
            start + 12,
            "Net Income",
            [
                "Depreciation & Amortization",
                "Share-Based Compensation",
                "Other Operating Activities",
            ],
            ["Operating Cash Flow"],
            True,
        )
        self.get_sum(
            start + 13,
            "Capital Expenditures",
            ["Acquisitions", "Change in Investments", "Other Investing Activities"],
            ["Investing Cash Flow"],
            True,
        )
        self.get_sum(
            start + 14,
            "Dividends Paid",
            [
                "Share Issuance / Repurchase",
                "Debt Issued / Paid",
                "Other Financing Activities",
            ],
            ["Financing Cash Flow"],
            True,
        )
        self.get_sum(
            start + 15,
            "Operating Cash Flow",
            ["Investing Cash Flow", "Financing Cash Flow"],
            ["Net Cash Flow"],
            True,
        )

    def get_linear(self, x_ind: str, y_ind: str, no_neg: bool = False):
        x_type = "IS" if x_ind in self.df_is.index else "BS"
        y_type = "IS" if y_ind in self.df_is.index else "BS"
        x_df = self.df_is if x_type == "IS" else self.df_bs
        y_df = self.df_is if y_type == "IS" else self.df_bs
        pre_x = (
            x_df.columns.to_numpy() if x_ind == "Date" else x_df.loc[x_ind].to_numpy()
        )

        vfunc = np.vectorize(string_float)
        pre_x = vfunc(pre_x)

        if x_ind == "Date":
            pre_x = pre_x - np.min(pre_x)
        x = pre_x.reshape((-1, 1))
        pre_y = y_df.loc[y_ind].to_numpy()
        y = vfunc(pre_y)
        model = LinearRegression().fit(x, y)
        r_sq = model.score(x, y)
        r = abs(math.sqrt(r_sq))

        if r > 0.9:
            strength = "very strong"
        elif r > 0.7:
            strength = "strong"
        elif r > 0.5:
            strength = "moderate"
        elif r > 0.3:
            strength = "weak"
        else:
            strength = "very weak"

        row1 = (
            y_df.index.get_loc(y_ind)
            + 1
            + (self.is_start if y_type == "IS" else self.bs_start)
        )

        col = self.len_pred + self.len_data + 3
        self.ws1[f"{var.letters[col]}{row1}"] = float(model.coef_)
        self.ws1[f"{var.letters[col+1]}{row1}"] = float(model.intercept_)
        self.ws1[f"{var.letters[col+2]}{row1}"] = var.letters[self.letter]
        self.ws1[f"{var.letters[col+2]}{row1}"].font = Font(color="FF0000")
        self.ws3[f"A{self.letter+4}"] = var.letters[self.letter]
        self.ws3[f"B{self.letter+4}"] = (
            f"The correlation between {x_ind.lower()} and {y_ind.lower()}"
            f" is {strength} with a correlation coefficient of {r:.4f}."
        )

        col = self.len_data + 1
        for i in range(self.len_pred):
            if x_ind == "Date":
                base = (
                    f"(({var.letters[col+i]}4-B4)*{var.letters[col+self.len_pred+2]}"
                    f"{row1})+{var.letters[col+self.len_pred+3]}{row1}"
                )
            else:
                row_n = (
                    x_df.index.get_loc(x_ind) + 1 + self.is_start
                    if x_type == "IS"
                    else self.bs_start
                )
                base = (
                    f"({var.letters[col+i]}{row_n}*{var.letters[col+self.len_pred+2]}{row1})"
                    f"+{var.letters[col+self.len_pred+3]}{row1}"
                )

            self.ws1[f"{var.letters[col+i]}{row1}"] = (
                f"=max({base},0)" if no_neg else f"={base}"
            )
            self.ws1[f"{var.letters[col+i]}{row1}"].number_format = var.fmt_acct

        self.letter += 1

    def get_sum(
        self,
        row: int_or_str,
        first: str,
        adds: List[str],
        subtracts: List[str],
        audit: bool = False,
        text: str = None,
    ):
        col = 1 if audit else self.len_data + 1
        for i in range(self.len_data if audit else self.len_pred):
            sum_formula = f"={var.letters[col+i]}{self.title_to_row(first)}"
            for item in adds:
                sum_formula += f"+{var.letters[col+i]}{self.title_to_row(item)}"
            for item in subtracts:
                sum_formula += f"-{var.letters[col+i]}{self.title_to_row(item)}"
            rowI = row if isinstance(row, int) else self.title_to_row(row)
            self.ws1[f"{var.letters[col+i]}{rowI}"] = sum_formula
            self.ws1[f"{var.letters[col+i]}{rowI}"].number_format = var.fmt_acct
        if text:
            self.custom_exp(row, text)

    def title_to_row(self, title: str):
        df = (
            self.df_is
            if title in self.df_is.index
            else self.df_bs
            if title in self.df_bs.index
            else self.df_cf
        )
        ind = (
            df.index.get_loc(title)
            + 1
            + (
                self.is_start
                if title in self.df_is.index
                else self.bs_start
                if title in self.df_bs.index
                else self.cf_start
            )
        )
        return ind

    def custom_exp(self, row: int_or_str, text: str, ws: int = 1, column: str = None):
        if ws == 1:
            rowT = row if isinstance(row, int) else self.title_to_row(row)
            col = self.len_pred + self.len_data + 3
            self.ws1[f"{var.letters[col+2]}{rowT}"] = var.letters[self.letter]
            self.ws1[f"{var.letters[col+2]}{rowT}"].font = Font(color="FF0000")
        if ws == 2:
            self.ws2[f"{column}{row}"] = var.letters[self.letter]
            self.ws2[f"{column}{row}"].font = Font(color="FF0000")

        self.ws3[f"A{self.letter+4}"] = var.letters[self.letter]
        self.ws3[f"B{self.letter+4}"] = text
        self.letter += 1
