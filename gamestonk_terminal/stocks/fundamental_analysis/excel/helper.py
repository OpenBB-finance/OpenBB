""" Excel Helper Functions """
__docformat__ = "numpy"

from typing import List, Union

import pandas as pd
from openpyxl import worksheet

opts = Union[int, str, float]


def string_float(string: str):
    """Numpy vectorize function to convert strings to floats"""
    return float(string.replace(",", ""))


def insert_row(name: str, index: str, df: pd.DataFrame, row_value: List[str]):
    """Allows a row to be inserted after a given row in a pandas dataframe"""
    pd.options.mode.chained_assignment = None
    if name not in df.index:
        row_number = df.index.get_loc(index) + 1
        df1 = df[0:row_number]
        df2 = df[row_number:]
        df1.loc[name] = row_value
        df_result = pd.concat([df1, df2])
        return df_result
    return df


def set_cell(
    ws: worksheet,
    cell: str,
    text: opts = None,
    font: str = None,
    border: str = None,
    fill: str = None,
    alignment: str = None,
    num_form: str = None,
):
    """Sets the value of the cell to given text and formats based on specified arguments"""
    ws[cell] = text
    if font:
        ws[cell].font = font
    if border:
        ws[cell].border = border
    if fill:
        ws[cell].fill = fill
    if alignment:
        ws[cell].alignment = alignment
    if num_form:
        ws[cell].number_format = num_form
