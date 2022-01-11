"""Yfinance options view"""
__docformat__ = "numpy"

import os
import math
from bisect import bisect_left
from typing import List, Dict, Any
from datetime import datetime, date, timedelta

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mtick
from scipy.stats import binom
import numpy as np
import pandas as pd
import seaborn as sns
import yfinance as yf
from tabulate import tabulate
from openpyxl import Workbook

import gamestonk_terminal.config_plot as cfp
import gamestonk_terminal.feature_flags as gtff
from gamestonk_terminal.helper_funcs import export_data, plot_autoscale
from gamestonk_terminal.stocks.options import op_helpers, yfinance_model
from gamestonk_terminal.stocks.options.yfinance_model import (
    generate_data,
    get_option_chain,
    get_price,
)
from gamestonk_terminal.helper_funcs import get_rf
from gamestonk_terminal.rich_config import console


def plot_oi(
    ticker: str,
    expiry: str,
    min_sp: float,
    max_sp: float,
    calls_only: bool,
    puts_only: bool,
    export: str = "",
):
    """Plot open interest

    Parameters
    ----------
    ticker: str
        Ticker
    expiry: str
        Expiry date for options
    min_sp: float
        Min strike to consider
    max_sp: float
        Max strike to consider
    calls_only: bool
        Show calls only
    puts_only: bool
        Show puts only
    export: str
        Format to export file
    """
    options = yfinance_model.get_option_chain(ticker, expiry)
    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "oi_yf",
        options,
    )
    calls = options.calls
    puts = options.puts
    current_price = float(yf.Ticker(ticker).info["regularMarketPrice"])

    if min_sp == -1:
        min_strike = 0.75 * current_price
    else:
        min_strike = min_sp

    if max_sp == -1:
        max_strike = 1.25 * current_price
    else:
        max_strike = max_sp

    if calls_only and puts_only:
        console.print("Both flags selected, please select one", "\n")
        return

    call_oi = calls.set_index("strike")["openInterest"] / 1000
    put_oi = puts.set_index("strike")["openInterest"] / 1000

    df_opt = pd.merge(call_oi, put_oi, left_index=True, right_index=True)
    df_opt = df_opt.rename(
        columns={"openInterest_x": "OI_call", "openInterest_y": "OI_put"}
    )

    max_pain = op_helpers.calculate_max_pain(df_opt)
    plt.style.use("classic")
    fig, ax = plt.subplots(figsize=plot_autoscale(), dpi=cfp.PLOT_DPI)

    if not calls_only:
        put_oi.plot(
            x="strike",
            y="openInterest",
            label="Puts",
            ax=ax,
            marker="o",
            ls="-",
            c="r",
        )
    if not puts_only:
        call_oi.plot(
            x="strike",
            y="openInterest",
            label="Calls",
            ax=ax,
            marker="o",
            ls="-",
            c="g",
        )
    ax.axvline(current_price, lw=2, c="k", ls="--", label="Current Price", alpha=0.7)
    ax.axvline(max_pain, lw=3, c="k", label=f"Max Pain: {max_pain}", alpha=0.7)
    ax.grid("on")
    ax.set_xlabel("Strike Price")
    ax.set_ylabel("Open Interest (1k) ")
    ax.set_xlim(min_strike, max_strike)

    if gtff.USE_ION:
        plt.ion()

    ax.set_title(f"Open Interest for {ticker.upper()} expiring {expiry}")
    plt.legend(loc=0)
    fig.tight_layout(pad=1)

    plt.show()
    console.print("")


def plot_vol(
    ticker: str,
    expiry: str,
    min_sp: float,
    max_sp: float,
    calls_only: bool,
    puts_only: bool,
    export: str = "",
):
    """Plot volume

    Parameters
    ----------
    ticker: str
        Ticker
    expiry: str
        Expiry date for options
    min_sp: float
        Min strike to consider
    max_sp: float
        Max strike to consider
    calls_only: bool
        Show calls only
    puts_only: bool
        Show puts only
    export: str
        Format to export file
    """
    options = yfinance_model.get_option_chain(ticker, expiry)
    calls = options.calls
    puts = options.puts
    current_price = float(yf.Ticker(ticker).info["regularMarketPrice"])

    if min_sp == -1:
        min_strike = 0.75 * current_price
    else:
        min_strike = min_sp

    if max_sp == -1:
        max_strike = 1.25 * current_price
    else:
        max_strike = max_sp

    if calls_only and puts_only:
        console.print("Both flags selected, please select one", "\n")
        return

    call_v = calls.set_index("strike")["volume"] / 1000
    put_v = puts.set_index("strike")["volume"] / 1000
    plt.style.use("classic")
    fig, ax = plt.subplots(figsize=plot_autoscale(), dpi=cfp.PLOT_DPI)

    if not calls_only:
        put_v.plot(
            x="strike",
            y="volume",
            label="Puts",
            ax=ax,
            marker="o",
            ls="-",
            c="r",
        )
    if not puts_only:
        call_v.plot(
            x="strike",
            y="volume",
            label="Calls",
            ax=ax,
            marker="o",
            ls="-",
            c="g",
        )
    ax.axvline(current_price, lw=2, c="k", ls="--", label="Current Price", alpha=0.7)
    ax.grid("on")
    ax.set_xlabel("Strike Price")
    ax.set_ylabel("Volume (1k) ")
    ax.set_xlim(min_strike, max_strike)

    if gtff.USE_ION:
        plt.ion()

    ax.set_title(f"Volume for {ticker.upper()} expiring {expiry}")
    plt.legend(loc=0)
    fig.tight_layout(pad=1)

    plt.show()
    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "vol_yf",
        options,
    )
    console.print("")


def plot_volume_open_interest(
    ticker: str,
    expiry: str,
    min_sp: float,
    max_sp: float,
    min_vol: float,
    export: str = "",
):
    """Plot volume and open interest

    Parameters
    ----------
    ticker: str
        Stock ticker
    expiry: str
        Option expiration
    min_sp: float
        Min strike price
    max_sp: float
        Max strike price
    min_vol: float
        Min volume to consider
    export: str
        Format for exporting data
    """

    options = yfinance_model.get_option_chain(ticker, expiry)
    calls = options.calls
    puts = options.puts
    current_price = float(yf.Ticker(ticker).info["regularMarketPrice"])

    # Process Calls Data
    df_calls = calls.pivot_table(
        index="strike", values=["volume", "openInterest"], aggfunc="sum"
    ).reindex()
    df_calls["strike"] = df_calls.index
    df_calls["type"] = "calls"
    df_calls["openInterest"] = df_calls["openInterest"]
    df_calls["volume"] = df_calls["volume"]
    df_calls["oi+v"] = df_calls["openInterest"] + df_calls["volume"]
    df_calls["spot"] = round(current_price, 2)

    df_puts = puts.pivot_table(
        index="strike", values=["volume", "openInterest"], aggfunc="sum"
    ).reindex()
    df_puts["strike"] = df_puts.index
    df_puts["type"] = "puts"
    df_puts["openInterest"] = df_puts["openInterest"]
    df_puts["volume"] = -df_puts["volume"]
    df_puts["openInterest"] = -df_puts["openInterest"]
    df_puts["oi+v"] = df_puts["openInterest"] + df_puts["volume"]
    df_puts["spot"] = round(current_price, 2)

    call_oi = calls.set_index("strike")["openInterest"] / 1000
    put_oi = puts.set_index("strike")["openInterest"] / 1000

    df_opt = pd.merge(call_oi, put_oi, left_index=True, right_index=True)
    df_opt = df_opt.rename(
        columns={"openInterest_x": "OI_call", "openInterest_y": "OI_put"}
    )

    max_pain = op_helpers.calculate_max_pain(df_opt)

    if min_vol == -1 and min_sp == -1 and max_sp == -1:
        # If no argument provided, we use the percentile 50 to get 50% of upper volume data
        volume_percentile_threshold = 50
        min_vol_calls = np.percentile(df_calls["oi+v"], volume_percentile_threshold)
        min_vol_puts = np.percentile(df_puts["oi+v"], volume_percentile_threshold)

        df_calls = df_calls[df_calls["oi+v"] > min_vol_calls]
        df_puts = df_puts[df_puts["oi+v"] < min_vol_puts]

    else:
        if min_vol > -1:
            df_calls = df_calls[df_calls["oi+v"] > min_vol]
            df_puts = df_puts[df_puts["oi+v"] < -min_vol]

        if min_sp > -1:
            df_calls = df_calls[df_calls["strike"] > min_sp]
            df_puts = df_puts[df_puts["strike"] > min_sp]

        if max_sp > -1:
            df_calls = df_calls[df_calls["strike"] < max_sp]
            df_puts = df_puts[df_puts["strike"] < max_sp]

    if df_calls.empty and df_puts.empty:
        console.print(
            "The filtering applied is too strong, there is no data available for such conditions.\n"
        )
        return

    # Initialize the matplotlib figure
    _, ax = plt.subplots(figsize=plot_autoscale(), dpi=cfp.PLOT_DPI)

    # make x axis symmetric
    axis_origin = max(abs(max(df_puts["oi+v"])), abs(max(df_calls["oi+v"])))
    ax.set_xlim(-axis_origin, +axis_origin)

    sns.set_style(style="darkgrid")

    g = sns.barplot(
        x="oi+v",
        y="strike",
        data=df_calls,
        label="Calls: Open Interest",
        color="lightgreen",
        orient="h",
    )

    g = sns.barplot(
        x="volume",
        y="strike",
        data=df_calls,
        label="Calls: Volume",
        color="green",
        orient="h",
    )

    g = sns.barplot(
        x="oi+v",
        y="strike",
        data=df_puts,
        label="Puts: Open Interest",
        color="pink",
        orient="h",
    )

    g = sns.barplot(
        x="volume",
        y="strike",
        data=df_puts,
        label="Puts: Volume",
        color="red",
        orient="h",
    )

    # draw spot line
    s = [float(strike.get_text()) for strike in ax.get_yticklabels()]
    spot_index = bisect_left(s, current_price)  # find where the spot is on the graph
    spot_line = ax.axhline(spot_index, ls="--", color="dodgerblue", alpha=0.3)

    # draw max pain line
    max_pain_index = bisect_left(s, max_pain)
    max_pain_line = ax.axhline(max_pain_index, ls="-", color="black", alpha=0.3)
    max_pain_line.set_linewidth(3)

    # format ticklabels without - for puts
    g.set_xticks(g.get_xticks())
    xlabels = [f"{x:,.0f}".replace("-", "") for x in g.get_xticks()]
    g.set_xticklabels(xlabels)

    plt.title(
        f"{ticker} volumes for {expiry} (open interest displayed only during market hours)"
    )
    ax.invert_yaxis()

    _ = ax.legend()
    handles, _ = ax.get_legend_handles_labels()
    handles.append(spot_line)
    handles.append(max_pain_line)

    # create legend labels + add to graph
    labels = [
        "Calls open interest",
        "Calls volume ",
        "Puts open interest",
        "Puts volume",
        "Current stock price",
        f"Max pain = {max_pain}",
    ]

    plt.legend(handles=handles[:], labels=labels)
    sns.despine(left=True, bottom=True)

    if gtff.USE_ION:
        plt.ion()
    plt.show()
    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "voi_yf",
        options,
    )
    console.print("")


def plot_plot(
    ticker: str, expiration: str, put: bool, x: str, y: str, custom: str, export: str
) -> None:
    """Generate a graph custom graph based on user input

    Parameters
    ----------
    ticker: str
        Stock ticker
    expiration: str
        Option expiration
    min_sp: float
        Min strike price
    put: bool
        put option instead of call
    x: str
        variable to display in x axis
    y: str
        variable to display in y axis
    custom: str
        type of plot
    export: str
        type of data to export
    """
    convert = {
        "ltd": "lastTradeDate",
        "s": "strike",
        "lp": "lastPrice",
        "b": "bid",
        "a": "ask",
        "c": "change",
        "pc": "percentChange",
        "v": "volume",
        "oi": "openInterest",
        "iv": "impliedVolatility",
    }

    x = convert[x]
    y = convert[y]
    varis = op_helpers.opt_chain_cols
    chain = yfinance_model.get_option_chain(ticker, expiration)
    values = chain.puts if put else chain.calls
    _, ax = plt.subplots()
    if custom == "smile":
        x = "strike"
        y = "impliedVolatility"
    x_data = values[x]
    y_data = values[y]
    ax.plot(x_data, y_data, "--bo")
    word = "puts" if put else "calls"
    ax.set_title(
        f"{varis[y]['label']} vs. {varis[x]['label']} for {ticker} {word} on {expiration}"
    )
    ax.set_ylabel(varis[y]["label"])
    ax.set_xlabel(varis[x]["label"])
    if varis[x]["format"] == "date":
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
        plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=1))
        plt.gcf().autofmt_xdate()
    elif varis[x]["format"]:
        ax.xaxis.set_major_formatter(varis[x]["format"])
    if varis[y]["format"] == "date":
        plt.gca().yaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
        plt.gca().yaxis.set_major_locator(mdates.DayLocator(interval=1))
    elif varis[y]["format"]:
        ax.yaxis.set_major_formatter(varis[y]["format"])
    if gtff.USE_ION:
        plt.ion()
    plt.show()
    export_data(export, os.path.dirname(os.path.abspath(__file__)), "plot")
    console.print("")


def plot_payoff(
    current_price: float,
    options: List[Dict[Any, Any]],
    underlying: int,
    ticker: str,
    expiration: str,
) -> None:
    """Generate a graph showing the option payoff diagram"""
    x, yb, ya = generate_data(current_price, options, underlying)
    _, ax = plt.subplots()
    if ya:
        ax.plot(x, yb, label="Payoff Before Premium")
        ax.plot(x, ya, label="Payoff After Premium")
    else:
        ax.plot(x, yb, label="Payoff")
    ax.set_title(f"Option Payoff Diagram for {ticker} on {expiration}")
    ax.set_ylabel("Profit")
    ax.set_xlabel("Underlying Asset Price at Expiration")
    ax.xaxis.set_major_formatter("${x:.2f}")
    ax.yaxis.set_major_formatter("${x:.2f}")
    plt.legend()
    if gtff.USE_ION:
        plt.ion()
    plt.show()
    console.print("")


def show_parity(
    ticker: str, exp: str, put: bool, ask: bool, mini: float, maxi: float, export: str
) -> None:
    """Prints options and whether they are under or over priced [Source: Yahoo Finance]

    Parameters
    ----------
    ticker : str
        Ticker to get expirations for
    exp : str
        Expiration to use for options
    put : bool
        Whether to use puts or calls
    ask : bool
        Whether to use ask or lastPrice
    mini : float
        Minimum strike price to show
    maxi : float
        Maximum strike price to show
    export : str
        Export data
    """
    r_date = datetime.strptime(exp, "%Y-%m-%d").date()
    delta = (r_date - date.today()).days
    rate = ((1 + get_rf()) ** (delta / 365)) - 1
    stock = get_price(ticker)

    div_info = yfinance_model.get_dividend(ticker)
    div_dts = div_info.index.values.tolist()

    if div_dts:
        last_div = pd.to_datetime(div_dts[-1])

        if len(div_dts) > 3:
            avg_div = np.mean(div_info.to_list()[-4:])
        else:
            avg_div = np.mean(div_info.to_list())

        next_div = last_div + timedelta(days=91)
        dividends = []
        while next_div < datetime.strptime(exp, "%Y-%m-%d"):
            day_dif = (next_div - datetime.now()).days
            dividends.append((avg_div, day_dif))
            next_div += timedelta(days=91)
        div_pvs = [x[0] / ((1 + get_rf()) ** (x[1] / 365)) for x in dividends]
        pv_dividend = sum(div_pvs)
    else:
        pv_dividend = 0

    chain = get_option_chain(ticker, exp)
    name = "ask" if ask else "lastPrice"
    o_type = "put" if put else "call"

    calls = chain.calls[["strike", name]].copy()
    calls = calls.rename(columns={name: "callPrice"})
    puts = chain.puts[["strike", name]].copy()
    puts = puts.rename(columns={name: "putPrice"})

    opts = pd.merge(calls, puts, on="strike")
    opts = opts.dropna()
    opts = opts.loc[opts["callPrice"] * opts["putPrice"] != 0]

    opts["callParity"] = (
        opts["putPrice"] + stock - (opts["strike"] / (1 + rate)) - pv_dividend
    )
    opts["putParity"] = (
        (opts["strike"] / (1 + rate)) + opts["callPrice"] - stock + pv_dividend
    )

    diff = o_type + " Difference"
    opts[diff] = opts[o_type + "Price"] - opts[o_type + "Parity"]
    opts["distance"] = abs(stock - opts["strike"])
    filtered = opts.copy()

    if mini is None:
        mini = filtered.strike.quantile(0.25)
    if maxi is None:
        maxi = filtered.strike.quantile(0.75)

    filtered = filtered.loc[filtered["strike"] >= mini]
    filtered = filtered.loc[filtered["strike"] <= maxi]

    show = filtered[["strike", diff]].copy()

    console.print("Warning: Low volume options may be difficult to trade.\n")
    if ask:
        console.print("Warning: Options with no current ask price not shown.\n")

    if gtff.USE_TABULATE_DF:
        print(
            tabulate(
                show,
                headers=[x.title() for x in show.columns],
                tablefmt="fancy_grid",
                showindex=False,
                floatfmt=".2f",
            )
        )
    else:
        console.print(show.to_string(index=False))

    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "parity",
        show,
    )
    console.print("")


def risk_neutral_vals(
    ticker: str,
    exp: str,
    put: bool,
    df: pd.DataFrame,
    mini: float,
    maxi: float,
    risk: float,
) -> None:
    """Prints current options prices and risk neutral values [Source: Yahoo Finance]

    Parameters
    ----------
    ticker : str
        Ticker to get expirations for
    exp : str
        Expiration to use for options
    put : bool
        Whether to use puts or calls
    df : pd.DataFrame
        Estimates for stocks prices and probabilities
    mini : float
        Minimum strike price to show
    maxi : float
        Maximum strike price to show
    risk : float
        The risk-free rate for the asset
    """
    if put:
        chain = get_option_chain(ticker, exp).puts
    else:
        chain = get_option_chain(ticker, exp).calls

    r_date = datetime.strptime(exp, "%Y-%m-%d").date()
    delta = (r_date - date.today()).days
    vals = []
    if risk is None:
        risk = get_rf()
    for _, row in chain.iterrows():
        vals.append(
            [
                row["strike"],
                row["lastPrice"],
                op_helpers.rn_payoff(row["strike"], df, put, delta, risk),
            ]
        )
    new_df = pd.DataFrame(vals, columns=["Strike", "Last Price", "Value"], dtype=float)
    new_df["Difference"] = new_df["Last Price"] - new_df["Value"]

    if mini is None:
        mini = new_df.Strike.quantile(0.25)
    if maxi is None:
        maxi = new_df.Strike.quantile(0.75)

    new_df = new_df[new_df["Strike"] >= mini]
    new_df = new_df[new_df["Strike"] <= maxi]

    if gtff.USE_TABULATE_DF:
        print(
            tabulate(
                new_df,
                headers=[x.title() for x in new_df.columns],
                tablefmt="fancy_grid",
                showindex=False,
                floatfmt=".2f",
            )
        )
    else:
        console.print(new_df.to_string(index=False))
    console.print("")


def plot_expected_prices(
    und_vals: List[List[float]], p: float, ticker: str, expiration: str
) -> None:
    """Plot expected prices of the underlying asset at expiration

    Parameters
    ----------
    und_vals : List[List[float]]
        The expected underlying values at the expiration date
    p : float
        The probability of the stock price moving upward each round
    ticker : str
        The ticker of the option's underlying asset
    expiration : str
        The expiration for the option
    """
    up_moves = list(range(len(und_vals[-1])))
    up_moves.reverse()
    probs = [binom.pmf(r, len(up_moves), p) for r in up_moves]
    fig, ax = plt.subplots(figsize=plot_autoscale(), dpi=cfp.PLOT_DPI)
    ax.set_title(f"Probabilities for ending prices of {ticker} on {expiration}")
    ax.xaxis.set_major_formatter("${x:1.2f}")
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    plt.plot(und_vals[-1], probs)
    fig.tight_layout()
    if gtff.USE_ION:
        plt.ion()
    plt.show()


def export_binomial_calcs(
    up: float,
    prob_up: float,
    discount: float,
    und_vals: List[List[float]],
    opt_vals: List[List[float]],
    days: int,
    ticker: str,
) -> None:
    """Create an excel spreadsheet with binomial tables for underlying asset value and option value

    Parameters
    ----------
    up : float
        The stock's increase on an upward move
    prob_up : float
        The probability of an upward move
    discount : float
        The daily discount rate
    und_vals : List[List[float]]
        The underlying asset values at each step
    opt_vals : List[List[float]]
        The values for the option at each step
    days : int
        The number of days until the option expires
    ticker : str
        The ticker for the company
    """
    letters = [
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
    ]
    letters += [
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",
    ]
    opts = (
        [f"{x}" for x in letters]
        + [f"{x}{y}" for x in letters for y in letters]
        + [f"{x}{y}{z}" for x in letters for y in letters for z in letters]
    )
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Up Move"
    ws["B1"] = up
    ws["A2"] = "Down Move"
    ws["B2"] = 1 / up
    ws["D1"] = "Prob Up"
    ws["E1"] = prob_up
    ws["D2"] = "Prob Down"
    ws["E2"] = 1 - prob_up
    ws["D3"] = "Discount"
    ws["E3"] = discount
    ws["A4"] = "Binomial Tree for Underlying Values"
    for i, _ in enumerate(und_vals):
        for j, _ in enumerate(und_vals[i]):
            ws[f"{opts[i]}{j+5}"] = und_vals[i][j]

    ws[f"A{days+7}"] = "Binomial Tree for Option Values"
    for i, _ in enumerate(opt_vals):
        for j, _ in enumerate(opt_vals[i]):
            ws[f"{opts[i]}{j+8+days}"] = opt_vals[i][j]

    trypath = os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..")),
        "exports",
        "stocks",
        "options",
        f"{ticker} {datetime.now()}.xlsx",
    )
    wb.save(trypath)
    console.print(
        f"Analysis ran for {ticker}\nPlease look in {trypath} for the file.\n"
    )


def show_binom(
    ticker: str,
    expiration: str,
    strike: float,
    put: bool,
    europe: bool,
    export: bool,
    plot: bool,
    vol: float,
) -> None:
    """Get binomial pricing for option

    Parameters
    ----------
    ticker : str
        The ticker of the option's underlying asset
    expiration : str
        The expiration for the option
    strike : float
        The strike price for the option
    put : bool
        Value a put instead of a call
    europe : bool
        Value a European option instead of an American option
    export : bool
        Export the options data to an excel spreadsheet
    plot : bool
        Show a graph of expected ending prices
    vol : float
        The annualized volatility for the underlying asset
    """
    # Base variables to calculate values
    info = yfinance_model.get_info(ticker)
    price = info["regularMarketPrice"]
    if vol is None:
        closings = yfinance_model.get_closing(ticker)
        vol = (closings / closings.shift()).std() * (252 ** 0.5)
    div_yield = (
        info["trailingAnnualDividendYield"]
        if info["trailingAnnualDividendYield"] is not None
        else 0
    )
    delta_t = 1 / 252
    rf = get_rf()
    exp_date = datetime.strptime(expiration, "%Y-%m-%d").date()
    today = date.today()
    days = (exp_date - today).days

    # Binomial pricing specific variables
    up = math.exp(vol * (delta_t ** 0.5))
    down = 1 / up
    prob_up = (math.exp((rf - div_yield) * delta_t) - down) / (up - down)
    prob_down = 1 - prob_up
    discount = math.exp(delta_t * rf)

    und_vals: List[List[float]] = [[price]]

    # Binomial tree for underlying values
    for i in range(days):
        cur_date = today + timedelta(days=i + 1)
        if cur_date.weekday() < 5:
            last = und_vals[-1]
            new = [x * up for x in last]
            new.append(last[-1] * down)
            und_vals.append(new)

    # Binomial tree for option values
    if put:
        opt_vals = [[max(strike - x, 0) for x in und_vals[-1]]]
    else:
        opt_vals = [[max(x - strike, 0) for x in und_vals[-1]]]

    j = 2
    while len(opt_vals[0]) > 1:
        new_vals = []
        for i in range(len(opt_vals[0]) - 1):
            if europe:
                value = (
                    opt_vals[0][i] * prob_up + opt_vals[0][i + 1] * prob_down
                ) / discount
            else:
                if put:
                    value = max(
                        (opt_vals[0][i] * prob_up + opt_vals[0][i + 1] * prob_down)
                        / discount,
                        strike - und_vals[-j][i],
                    )
                else:
                    value = max(
                        (opt_vals[0][i] * prob_up + opt_vals[0][i + 1] * prob_down)
                        / discount,
                        und_vals[-j][i] - strike,
                    )
            new_vals.append(value)
        opt_vals.insert(0, new_vals)
        j += 1

    if export:
        export_binomial_calcs(up, prob_up, discount, und_vals, opt_vals, days, ticker)

    if plot:
        plot_expected_prices(und_vals, prob_up, ticker, expiration)

    console.print(
        f"{ticker} {'put' if put else 'call'} at ${strike:.2f} expiring on {expiration} is worth ${opt_vals[0][0]:.2f}\n"
    )


def display_vol_surface(ticker: str, export: str = ""):
    """Display vol surface

    Parameters
    ----------
    ticker: str
        Ticker to get surface for
    export: str
        Format to export data

    """
    data = yfinance_model.get_iv_surface(ticker)
    if data.empty:
        print(f"No options data found for {ticker}.\n")
        return
    X = data.dte
    Y = data.strike
    Z = data.impliedVolatility
    fig = plt.figure()
    ax = plt.axes(projection="3d")
    ax.plot_trisurf(X, Y, Z, cmap="jet", linewidth=0.2)
    ax.set_xlabel("DTE")
    ax.set_ylabel("Strike")
    ax.set_zlabel("IV")
    fig.tight_layout()
    fig.suptitle(f"Volatility Surface for {ticker}")
    if gtff.USE_ION:
        plt.ion()
    plt.show()
    export_data(
        export,
        os.path.dirname(os.path.abspath(__file__)),
        "vsurf",
        data,
    )
    print("")
